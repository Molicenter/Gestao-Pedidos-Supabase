import streamlit as st
import pandas as pd
import io
import time
import requests
from datetime import date, datetime, timezone, timedelta
from supabase import create_client, Client

# ─────────────────────────────────────────────────────────────────────────────
# ⚙️ CONSTANTES, ESTILOS E CONEXÕES GLOBAIS
# ─────────────────────────────────────────────────────────────────────────────
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

LOJAS_NOMES = ["Loja 01", "Loja 02", "Loja 03", "Loja 04", "Loja 05", "Loja 06", "Loja 07", "Loja 08"]

# ─────────────────────────────────────────────────────────────────────────────
# 🚚 MAPA FIXO FORNECEDOR → PRODUTOS (por código Iceasa) — usado só no FLV.
# Não vai pro banco: a visão "Pedido por Fornecedor" só PUXA quantidade (pedidos)
# e preço (Separação) já existentes. Pra mudar permanentemente quem vende o quê,
# edita-se este dicionário. Os "de linha" saem no layout girado (lojas nas linhas).
# ─────────────────────────────────────────────────────────────────────────────
MAPA_FORNECEDORES_FLV = {
    "NIDE": [45, 49, 67, 57, 46, 48, 47], "Claudir Mendes": [57, 46, 49], "SANDRO": [75],
    "DENIZE": [45, 49, 67, 57, 46, 48, 47, 41, 52, 69, 56], "JOVANO": [1746, 88, 49, 140, 85],
    "JEFINHO": [85, 140, 256, 267, 88, 57, 45, 49, 1662, 46], "LUCIANO": [61, 41, 49, 56, 45],
    "THIAGO": [61, 91, 67, 74, 49, 52, 45, 56], "CRISTIAN": [40, 949, 42, 83, 68, 538, 78],
    "ROGERIO NARANTE": [538], "FERNANDO NARANTE": [46], "SILVIO MAND SALSA": [76],
    "HORTA": [108, 109], "GLAUCIA MACIEL": [84, 85], "ALEMÃO": [39], "RENAN SS": [72],
    "NEGUIN": [85, 86, 88, 61, 45], "RODRIGO CHANAN": [85, 86, 88, 61, 1662, 140],
    "MARCELO MORANGO": [58], "JOÃO BATISTA": [79, 60, 56, 1662, 69], "GIACOMELLO": [95],
    "PRIMO": [240, 86, 49, 45, 88, 85], "RENATO MANDIOCA": [75], "THIAGO SERRA": [91, 49, 45, 56, 61],
    "TICO": [236, 237, 707, 2730, 42, 581, 78, 546, 80, 83, 540, 949, 40, 110, 68, 109],
    "ALGACIR": [1516, 53], "MAURICIO": [62], "PAULO IGASHIBAHI": [47, 48],
    "GILSOM BATATA": [508, 551], "DORI BATATA": [508, 551], "BANANA SANTOME": [2567, 2569, 2568],
    "MELANCIA CARLIN": [1], "MELANCIA MARCINHO": [673, 1, 3003], "RODRIGO BATATA": [508],
}
FORNECEDORES_LINHA_FLV = {"BANANA SANTOME", "MELANCIA CARLIN", "MELANCIA MARCINHO", "RODRIGO BATATA"}

@st.cache_resource
def obter_supabase() -> Client:
    # cache_resource: reaproveita o MESMO cliente entre reruns/sessões em vez de
    # criar uma conexão nova a cada clique. Ganho direto de velocidade.
    return create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])

def data_brasilia() -> date:
    # Data no fuso de Brasília (UTC-3 fixo). O servidor roda em UTC, então sem isso
    # um pedido lançado à noite cairia no dia seguinte.
    return (datetime.now(timezone.utc) - timedelta(hours=3)).date()

def data_hora_brasilia() -> str:
    # O servidor do Streamlit Cloud roda em UTC. Como Londrina/Brasília é UTC-3 fixo
    # (sem horário de verão desde 2019), basta subtrair 3h — sem depender de tzdata.
    agora = datetime.now(timezone.utc) - timedelta(hours=3)
    return agora.strftime("%d/%m/%Y %H:%M")

conn_pg = st.connection("banco_erp", type="sql")

# ─────────────────────────────────────────────────────────────────────────────
# 🔍 FUNÇÕES AUXILIARES DE CONSULTA E UTILITÁRIOS
# ─────────────────────────────────────────────────────────────────────────────
def converter_para_int_seguro(valor) -> int:
    if pd.isna(valor) or valor is None:
        return 0
    val_str = str(valor).strip().replace(',', '.')
    if val_str == "" or val_str.lower() in ["<na>", "none", "nan", "null", "-"]:
        return 0
    try:
        qtd_float = float(val_str)
        return int(qtd_float)
    except ValueError:
        return 0

def setor_usa_iceasa(setor) -> bool:
    # Cód. Iceasa só existe p/ FLV Normal e FLV Ofertas
    return str(setor).strip().lower() in ("flv normal", "flv ofertas")

def _normaliza_setor(setor) -> str:
    s = str(setor).strip().lower()
    for a, b in (("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),("â","a"),("ê","e"),("ô","o"),("ã","a"),("õ","o"),("ç","c")):
        s = s.replace(a, b)
    return s

def setor_pedido_texto(setor) -> bool:
    # Matéria Prima, Embalagem(ns) e Padaria/Confeitaria pedem em texto (fardo, cx, fd...).
    # Esses 3 setores saem SEM Total em tudo (telas, Excel e impressão).
    # "embalag" pega tanto "Embalagem" quanto "Embalagens".
    return any(k in _normaliza_setor(setor) for k in ("materia prima", "embalag", "padaria", "confeitaria"))

def setor_usa_obs_loja(setor) -> bool:
    # Embalagem(ns) e Padaria/Confeitaria têm campo "Observação Geral da Loja"
    # (a loja escreve; o admin vê na Separação).
    return any(k in _normaliza_setor(setor) for k in ("embalag", "padaria", "confeitaria"))

def setor_usa_sem_pedido(setor) -> bool:
    # Botão "Sem Pedido Hoje" (zera o pedido da loja + avisa no Telegram) só aparece
    # no Açougue Adriano e no Pioneiro+BF+Paraná (que no Supabase é "Açougue Especiais").
    return _normaliza_setor(setor) in ("acougue adriano", "acougue especiais")

def setor_eh_pecas_manoel(setor) -> bool:
    # Peças Açougue - Manoel: NÃO tem código ERP, nem estoque do ERP, nem média 90d.
    return "manoel" in _normaliza_setor(setor)

def setor_usa_erp(setor) -> bool:
    # Setores ligados ao ERP (têm Cód. ERP e puxam estoque). Peças Manoel não usa.
    return not setor_eh_pecas_manoel(setor)

def grupo_fornecedor(forn) -> str:
    # Agrupa os fornecedores "BIG FRANGO - *" sob um único rótulo de filtro "BIG FRANGO".
    # A coluna Fornecedor continua mostrando o nome completo de cada tipo (Mix Balcão etc.).
    if _normaliza_setor(forn).startswith("big frango"):
        return "BIG FRANGO"
    return str(forn).strip()

# Views de média no ERP (período base de 90 dias)
VIEWS_MEDIA = {
    "Média Semanal": "python_90dSEMANA",
    "Diária": "python_90dDIARIA",
    "Seg-Ter": "python_90dSEGTER",
    "Qua-Qui": "python_90dQUAQUI",
    "Sex-Sab-Dom": "python_90dSEXSABDOM",
}

def setor_usa_media(setor) -> bool:
    # Embalagem/Padaria/Confeitaria/Matéria Prima e Peças Açougue-Manoel NÃO usam a coluna Média.
    return not setor_pedido_texto(setor) and not setor_eh_pecas_manoel(setor)

def filtro_media_padrao(setor) -> str:
    # Filtro de média já pré-selecionado por setor (o admin ainda pode trocar).
    s = _normaliza_setor(setor)
    if "folhagem" in s:
        return "Diária"
    if "flv normal" in s:
        # Quinta-feira (pedido p/ Sex-Sáb-Dom) → Sex-Sab-Dom; demais dias (inclui terça) → Qua-Qui
        return "Sex-Sab-Dom" if data_brasilia().weekday() == 3 else "Qua-Qui"
    if "flv oferta" in s:
        return "Seg-Ter"
    # FLV Oriental, Açougue (Adriano/Pioneiro/BF/Paraná) e demais → Média Semanal
    return "Média Semanal"

def setor_eh_materia_prima(setor) -> bool:
    # Só a Matéria Prima ganha a coluna Observação — e apenas na exportação Excel.
    return "materia prima" in _normaliza_setor(setor)

def iceasa_para_impressao(v) -> str:
    # Regra da impressão/Excel BOX: vazio se ausente OU acima de 9000; senão o número
    try:
        n = int(float(str(v).strip()))
    except (ValueError, TypeError):
        return ""
    return "" if n > 9000 else str(n)

def preco_para_celula(v) -> str:
    # float do banco -> texto BR com R$ p/ exibir no editor ("R$ 12,50"); vazio fica vazio (sem "None")
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    try:
        return f"R$ {float(v):.2f}".replace(".", ",")
    except (ValueError, TypeError):
        return ""

def celula_para_preco(s):
    # texto digitado ("R$ 12,50", "12.50", "12,5") -> float; vazio/inválido -> None
    if s is None:
        return None
    t = str(s).strip().replace("R$", "").replace(" ", "").replace(",", ".")
    if t == "":
        return None
    try:
        return round(float(t), 2)
    except ValueError:
        return None

def buscar_estoque_erp(loja_nome, codigos_erp, setor):
    if not codigos_erp: 
        return pd.DataFrame(columns=["Código", "Estoque"])
        
    loja_id = int(loja_nome.split()[-1])
    loja_id_str = f"{loja_id:03d}" 
    cods_str = ", ".join(map(str, set(codigos_erp)))
    
    coluna_alvo = "estoqueemb" if setor == "Embalagem" else "estoque"
    
    query = f"""
        SELECT cade_codigo AS "Código", {coluna_alvo} AS "Estoque"
        FROM python_estoque WHERE cade_codempresa = '{loja_id_str}' AND cade_codigo IN ({cods_str})
    """
    try: 
        return conn_pg.query(query, ttl=30) 
    except Exception as e: 
        st.error(f"Erro ao buscar estoque: {e}")
        return pd.DataFrame({"Código": codigos_erp, "Estoque": 0})

# ─────────────────────────────────────────────────────────────────────────────
# 📊 MOTORES DE EXPORTAÇÃO EXCEL CUSTOMIZADOS (OPENPYXL)
# ─────────────────────────────────────────────────────────────────────────────
def gerar_excel_download(df: pd.DataFrame, nome_aba: str, com_obs: bool = False, obs_rodape: str = "") -> bytes:
    df_export = df.copy()
    df_export = df_export.rename(columns={
        "Cód. ERP": "Código",
        "TOTAL GERAL": "Total",
        "Qtde Pedida": "Pedido",
        "Estoque ERP": "Estoque",
        "Observação": "Observação:",
    })

    # Se a tela já trouxe a Observação preenchida, ela vira a coluna do Excel.
    # Senão, com_obs cria a coluna em branco para preenchimento manual.
    if com_obs and not any("OBSERV" in str(c).upper() for c in df_export.columns):
        df_export["Observação:"] = ""

    def limpar_valor_excel(v):
        if pd.isna(v): return None
        v_str = str(v).strip()
        if v_str in ["", "-", "0", "0.0"]: return None
        try:
            f = float(v_str)
            if f == 0: return None
            return int(f) if f.is_integer() else f
        except ValueError:
            return v 

    for col in df_export.columns:
        col_upper = str(col).upper()
        if any(x in col_upper for x in ["LOJA", "PEDIDO", "ESTOQUE", "TOTAL", "MÉDIA"]):
            df_export[col] = df_export[col].apply(limpar_valor_excel)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name=nome_aba[:30], startrow=1)
        worksheet = writer.sheets[nome_aba[:30]]

        fill_header = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
        fill_green = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        font_header = Font(color="FFFFFF", bold=True)
        font_bold = Font(bold=True)
        
        border_thin = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
                             top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
                             
        border_header = Border(left=Side(style='dashed', color='FFFFFF'), right=Side(style='dashed', color='FFFFFF'),
                               top=Side(style='dashed', color='FFFFFF'), bottom=Side(style='dashed', color='FFFFFF'))
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        hoje_str = data_brasilia().strftime("%d/%m/%Y")
        worksheet["A1"] = f"Pedidos do dia {hoje_str}"
        worksheet["A1"].font = font_bold

        colunas_verdes = []
        col_total_idx = None
        cols_para_soma = []
        contador_loja = 0 

        for col_num, column_title in enumerate(df_export.columns, 1):
            col_name = str(column_title).upper()
            if "LOJA" in col_name:
                cols_para_soma.append(col_num)
                contador_loja += 1
                if contador_loja % 2 != 0:
                    colunas_verdes.append(col_num)
            elif any(x in col_name for x in ["PEDIDO", "TOTAL", "MÉDIA", "ESTOQUE"]):
                colunas_verdes.append(col_num)
            if "TOTAL" in col_name:
                col_total_idx = col_num

        for col_num, cell in enumerate(worksheet[2], 1):
            cell.fill = fill_header
            cell.font = font_header
            cell.border = border_header
            cell.alignment = align_center

        for row_num, row in enumerate(worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column), 3):
            for cell in row:
                cell.border = border_thin
                cell.font = font_bold
                nome_col_atual = str(df_export.columns[cell.column - 1]).upper()
                if any(x in nome_col_atual for x in ["FORNECEDOR", "DESCRIÇÃO", "PRODUTO"]):
                    cell.alignment = align_left
                else:
                    cell.alignment = align_center
                if cell.column in colunas_verdes:
                    cell.fill = fill_green

        if col_total_idx and cols_para_soma:
            letra_total = get_column_letter(col_total_idx)
            letra_primeira = get_column_letter(min(cols_para_soma))
            letra_ultima = get_column_letter(max(cols_para_soma))
            for row_num in range(3, worksheet.max_row + 1):
                worksheet[f"{letra_total}{row_num}"].value = f'=IF(SUM({letra_primeira}{row_num}:{letra_ultima}{row_num})>0, SUM({letra_primeira}{row_num}:{letra_ultima}{row_num}), "")'

        for col_num, column_title in enumerate(df_export.columns, 1):
            letra = get_column_letter(col_num)
            col_name = str(column_title).upper()
            if "FORNECEDOR" in col_name: 
                worksheet.column_dimensions[letra].width = 18
            elif "DESCRI" in col_name or "PRODUTO" in col_name: 
                worksheet.column_dimensions[letra].width = 45
            elif "CÓDIGO" in col_name: 
                worksheet.column_dimensions[letra].width = 12
            elif "MÉDIA" in col_name or "ESTOQUE" in col_name: 
                worksheet.column_dimensions[letra].width = 12
            elif "OBSERV" in col_name:
                worksheet.column_dimensions[letra].width = 22
            else: 
                worksheet.column_dimensions[letra].width = 9 

        # 📝 Observação Geral da Loja: escrita abaixo da tabela (só no pedido da loja)
        if obs_rodape and str(obs_rodape).strip():
            ncols = max(1, len(df_export.columns))
            r_lbl = worksheet.max_row + 2
            c_lbl = worksheet.cell(row=r_lbl, column=1, value="📝 Observação Geral da Loja:")
            c_lbl.font = font_bold
            if ncols > 1:
                worksheet.merge_cells(start_row=r_lbl, start_column=1, end_row=r_lbl, end_column=ncols)
            r_txt = r_lbl + 1
            c_txt = worksheet.cell(row=r_txt, column=1, value=str(obs_rodape).strip())
            c_txt.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c_txt.border = border_thin
            if ncols > 1:
                worksheet.merge_cells(start_row=r_txt, start_column=1, end_row=r_txt + 3, end_column=ncols)
            worksheet.row_dimensions[r_txt].height = 60

        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0 
        worksheet.page_margins.left = 1 / 2.54
        worksheet.page_margins.right = 1 / 2.54
        worksheet.page_margins.top = 1 / 2.54
        worksheet.page_margins.bottom = 1 / 2.54
        worksheet.page_margins.header = 0.5 / 2.54
        worksheet.page_margins.footer = 0.5 / 2.54

    return output.getvalue()

def gerar_excel_fornecedores(df: pd.DataFrame, nome_aba: str, sem_total: bool = False, com_obs: bool = False) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        worksheet = writer.book.create_sheet(nome_aba[:30])
        writer.book.active = worksheet
        
        fill_header = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
        fill_green = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        font_header = Font(color="FFFFFF", bold=True)
        font_bold = Font(bold=True)
        font_normal = Font(bold=False)
        border_thin = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
                             top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
                             
        border_header = Border(left=Side(style='dashed', color='FFFFFF'), right=Side(style='dashed', color='FFFFFF'),
                               top=Side(style='dashed', color='FFFFFF'), bottom=Side(style='dashed', color='FFFFFF'))
                               
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        def limpar(v):
            if pd.isna(v): return ""
            v_str = str(v).strip()
            if v_str in ["", "-", "0", "0.0", "nan"]: return ""
            try:
                f = float(v_str)
                if f == 0: return ""
                return int(f) if f.is_integer() else f
            except ValueError:
                return v_str

        # Layout: lojas vão das colunas 3..10. Depois (opcional) Total e/ou Observação.
        total_col = None if sem_total else 11
        if com_obs:
            obs_col = (total_col + 1) if total_col else 11
        else:
            obs_col = None
        last_col = 10
        for c in (total_col, obs_col):
            if c:
                last_col = max(last_col, c)

        worksheet.column_dimensions['A'].width = 12  
        worksheet.column_dimensions['B'].width = 45  
        for col_idx in range(3, last_col + 1): 
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 9
        if obs_col:
            worksheet.column_dimensions[get_column_letter(obs_col)].width = 22

        hoje_str = data_brasilia().strftime("%d/%m/%Y")
        worksheet["A1"] = "Tipo"
        worksheet["B1"] = "Descrição"
        worksheet["C1"] = f"Pedidos do dia {hoje_str}"
        worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=last_col) 
        
        for col_idx in range(1, last_col + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = fill_header
            cell.font = font_header
            cell.border = border_header
            cell.alignment = align_center

        current_row = 2
        lojas_cols = ["Loja 01", "Loja 02", "Loja 03", "Loja 04", "Loja 05", "Loja 06", "Loja 07", "Loja 08"]
        
        forns = sorted(df["fornecedor"].dropna().unique())
        for forn in forns:
            df_f = df[df["fornecedor"] == forn]
            
            worksheet.cell(row=current_row, column=1).value = ""
            worksheet.cell(row=current_row, column=2).value = forn
            
            for i, loja in enumerate(lojas_cols): 
                worksheet.cell(row=current_row, column=3+i).value = loja
                
            if total_col:
                worksheet.cell(row=current_row, column=total_col).value = "TOTAL"
            if obs_col:
                worksheet.cell(row=current_row, column=obs_col).value = "Observação"
            
            for col_idx in range(1, last_col + 1):
                cell = worksheet.cell(row=current_row, column=col_idx)
                cell.fill = fill_header
                cell.font = font_header
                cell.border = border_header
                cell.alignment = align_center
            
            current_row += 1
            
            for _, r in df_f.iterrows():
                c1 = worksheet.cell(row=current_row, column=1, value=limpar(r.get("Cód. ERP", "")))
                c1.alignment = align_center
                c1.border = border_thin
                c1.font = font_normal
                
                c2 = worksheet.cell(row=current_row, column=2, value=limpar(r.get("Produto", "")))
                c2.alignment = align_left
                c2.border = border_thin
                c2.font = font_normal
                
                for i, loja in enumerate(lojas_cols):
                    c_loja = worksheet.cell(row=current_row, column=3+i, value=limpar(r.get(loja, "")))
                    c_loja.alignment = align_center
                    c_loja.border = border_thin
                    c_loja.font = font_bold
                    if i % 2 == 0: 
                        c_loja.fill = fill_green
                
                if total_col:
                    c_total = worksheet.cell(row=current_row, column=total_col)
                    c_total.value = f'=IF(SUM(C{current_row}:J{current_row})>0, SUM(C{current_row}:J{current_row}), "")'
                    c_total.alignment = align_center
                    c_total.border = border_thin
                    c_total.font = font_bold
                    c_total.fill = fill_green
                if obs_col:
                    obs_v = r.get("Observação", "")
                    if obs_v is None or (isinstance(obs_v, float) and pd.isna(obs_v)):
                        obs_v = ""
                    obs_v = str(obs_v).strip()
                    if obs_v in ("nan", "None", "<NA>"):
                        obs_v = ""
                    c_obs = worksheet.cell(row=current_row, column=obs_col, value=obs_v)
                    c_obs.alignment = align_left
                    c_obs.border = border_thin
                    c_obs.font = font_normal
                current_row += 1
                
        if 'Sheet' in writer.book.sheetnames: 
            writer.book.remove(writer.book['Sheet'])

        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0 
        worksheet.page_margins.left = 1 / 2.54
        worksheet.page_margins.right = 1 / 2.54
        worksheet.page_margins.top = 1 / 2.54
        worksheet.page_margins.bottom = 1 / 2.54
        worksheet.page_margins.header = 0.5 / 2.54
        worksheet.page_margins.footer = 0.5 / 2.54

    return output.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# 📦 EXCEL "PEDIDO BOX" (padrão Molicenter) — exclusivo p/ FLV Normal e FLV Ofertas
# ─────────────────────────────────────────────────────────────────────────────
def gerar_excel_box(df: pd.DataFrame) -> bytes:
    """Layout 'PEDIDO BOX' replicando o molicenter_final.xlsx.
    A CODIGO (= Cód. Iceasa) | B PRODUTOS MOLICENTER | C-J 291..298 | K-P spacer oculto |
    Q TOTAL | R PREÇO | S OBS. Iceasa vazio ou > 9000 sai em branco na coluna A."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PEDIDO BOX"

    # paleta exata do template
    C_LARANJA = "C55A11"   # cabeçalho
    C_VERDE   = "E2EFDA"   # lojas ímpares (banda verde)
    C_BRANCO  = "FFFFFF"   # lojas pares
    C_VERDE_T = "C6EFCE"   # coluna TOTAL
    C_SALMAO  = "FCE4D6"   # coluna PREÇO
    FMT_REAL  = '[$R$-416]\\ #,##0.00'

    fonte = Font(name="Arial", size=9, bold=True)
    fonte_h = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    borda = Border(*([Side(style="thin", color="000000")] * 4))
    cc = Alignment(horizontal="center", vertical="center")
    cl = Alignment(horizontal="left", vertical="center")

    headers = ["CODIGO", "PRODUTOS MOLICENTER",
               "291", "292", "293", "294", "295", "296", "297", "298",
               "", "", "", "", "", "", "TOTAL", "PREÇO", "OBS:"]
    for col, txt in enumerate(headers, 1):
        cel = ws.cell(row=2, column=col, value=txt)
        cel.fill = PatternFill("solid", fgColor=C_LARANJA)
        cel.font = fonte_h
        cel.border = borda
        cel.alignment = cl if col == 2 else cc

    # fill de fundo de cada coluna de DADO
    fill_dados = {}
    for i in range(8):                       # C..J = lojas
        fill_dados[3 + i] = C_VERDE if i % 2 == 0 else C_BRANCO
    fill_dados[17] = C_VERDE_T               # Q TOTAL
    fill_dados[18] = C_SALMAO                # R PREÇO

    def _cod_box(v):
        try:
            n = int(float(str(v).strip()))
        except (ValueError, TypeError):
            return None
        return None if n > 9000 else n      # >9000 → célula vazia

    def _qtd_box(v):
        if v is None:
            return None
        s = str(v).strip()
        if s in ("", "-", "0", "0.0", "nan", "None", "<NA>"):
            return None
        try:
            f = float(s.replace(",", "."))
            return None if f == 0 else (int(f) if f.is_integer() else f)
        except ValueError:
            return None

    # alfabético por produto (igual ao template)
    df = df.copy()
    if "Descrição" in df.columns:
        df = df.sort_values(by="Descrição", key=lambda s: s.astype(str).str.lower())

    r = 3
    for _, row in df.iterrows():
        ws.cell(row=r, column=1, value=_cod_box(row.get("Cód. Iceasa")))
        ws.cell(row=r, column=2, value=str(row.get("Descrição", "")).strip())
        for i, loja in enumerate(LOJAS_NOMES):
            ws.cell(row=r, column=3 + i, value=_qtd_box(row.get(loja)))
        ws.cell(row=r, column=17, value=f'=IF(SUM(C{r}:J{r})=0,"",SUM(C{r}:J{r}))')
        for col in range(1, 20):
            cel = ws.cell(row=r, column=col)
            cel.font = fonte
            cel.border = borda
            cel.alignment = cl if col == 2 else cc
            cel.fill = PatternFill("solid", fgColor=fill_dados.get(col, C_BRANCO))
            if col == 18:
                cel.number_format = FMT_REAL    # R$ ao digitar
        r += 1

    larguras = {"A": 9, "B": 34, "C": 8.5, "D": 8.5, "E": 8.5, "F": 8.5, "G": 8.5,
                "H": 8.5, "I": 8.5, "J": 8.5, "K": 3, "L": 13, "M": 13, "N": 13,
                "O": 13, "P": 13, "Q": 8.5, "R": 12, "S": 22}
    for L, w in larguras.items():
        ws.column_dimensions[L].width = w
    for L in ["K", "L", "M", "N", "O", "P"]:     # 6 spacers ocultos
        ws.column_dimensions[L].hidden = True

    ws.row_dimensions[1].height = 6              # linha 1 reduzida
    ws.row_dimensions[2].height = 18
    for rr in range(3, r):
        ws.row_dimensions[rr].height = 15

    ws.auto_filter.ref = "A2:S2"
    ws.freeze_panes = "A3"

    # impressão RETRATO A4 (pedido do usuário)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.75
    ws.page_margins.right = 0.75
    ws.page_margins.top = 1.0
    ws.page_margins.bottom = 1.0

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# 🖨️ OUTROS UTILS DA INTERFACE
# ─────────────────────────────────────────────────────────────────────────────
def notificar_telegram(mensagem: str) -> bool:
    # Envia um aviso a um grupo do Telegram. Lê as credenciais de st.secrets["telegram"]
    # (bot_token e chat_id). Retorna True se o Telegram aceitou a mensagem.
    try:
        bot_token = st.secrets["telegram"]["bot_token"]
        chat_id = st.secrets["telegram"]["chat_id"]
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        resposta = requests.post(url, json={"chat_id": chat_id, "text": mensagem}, timeout=10)
        if resposta.status_code == 200:
            return True
        st.error(f"🚨 Erro do Telegram: {resposta.text}")
        return False
    except KeyError:
        st.error("⚠️ Credenciais do Telegram não configuradas no secrets ([telegram]).")
        return False
    except Exception as e:
        st.error(f"⚠️ Erro de conexão ao enviar o Telegram: {e}")
        return False

def injetar_botao_impressao():
    st.components.v1.html(
        """
        <button onclick="window.parent.print()" style="
            width: 100%;
            background-color: #f0f2f6;
            color: #31333f;
            border: 1px solid rgba(49, 51, 63, 0.2);
            padding: 0.53rem 1rem;
            border-radius: 0.5rem;
            cursor: pointer;
            font-weight: 500;
            font-size: 14px;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 8px;
            box-sizing: border-box;
            height: 38px;
        ">
            🖨️ Imprimir
        </button>
        """,
        height=42,
    )

def exibir_status_digitacao_lojas(df_pedidos_hoje, lojas_sem_pedido=None):
    st.markdown(
        f"<div class='no-print' style='font-weight:600; font-size:1.05rem; margin:2px 0 6px 0;'>"
        f"🏪 Status de Digitação das Lojas — {data_hora_brasilia()}</div>",
        unsafe_allow_html=True,
    )
    lojas_que_digitam = set()
    if not df_pedidos_hoje.empty:
        lojas_codigos = df_pedidos_hoje["loja"].unique()
        for c in lojas_codigos: 
            lojas_que_digitam.add(f"Loja {c:02d}")

    # Lojas que declararam "Sem Pedido Hoje" — contam como respondidas (ficam verdes).
    sem_pedido_nomes = {f"Loja {int(n):02d}" for n in (lojas_sem_pedido or set())}

    cols = st.columns(8)
    for i, loja_nome in enumerate(LOJAS_NOMES):
        with cols[i]:
            if loja_nome in lojas_que_digitam:
                st.markdown(f"<div class='no-print' style='text-align:center; background-color:#d4edda; color:#155724; padding:5px; border-radius:5px; font-size:11px; font-weight:bold;'>{loja_nome}<br>✅ OK</div>", unsafe_allow_html=True)
            elif loja_nome in sem_pedido_nomes:
                # Verde também, mas sinalizando que foi uma declaração de "sem pedido"
                st.markdown(f"<div class='no-print' style='text-align:center; background-color:#d4edda; color:#155724; padding:5px; border-radius:5px; font-size:11px; font-weight:bold;'>{loja_nome}<br>✅ Sem Pedido</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div class='no-print' style='text-align:center; background-color:#f8d7da; color:#721c24; padding:5px; border-radius:5px; font-size:11px; font-weight:bold;'>{loja_nome}<br>❌ Faltando</div>", unsafe_allow_html=True)
    st.markdown("<div class='no-print'><br></div>", unsafe_allow_html=True)

@st.cache_data(ttl=300, show_spinner=False)
def buscar_permissoes_setor(_supabase_client, codigos_setor, num_loja=None):
    # _supabase_client com underscore: o Streamlit NÃO tenta "hashear" o cliente.
    # A chave do cache é (codigos_setor, num_loja). TTL de 5 min; invalidado na hora
    # via st.cache_data.clear() quando o catálogo/permissões mudam.
    if not codigos_setor: 
        return pd.DataFrame(columns=["codigo_produto", "loja", "disponivel"])
    dfs = []
    for i in range(0, len(codigos_setor), 200):
        lote = codigos_setor[i:i+200]
        query = _supabase_client.table("produtos_lojas").select("codigo_produto, loja, disponivel").in_("codigo_produto", lote)
        if num_loja is not None: 
            query = query.eq("loja", num_loja)
        resp = query.execute()
        if resp.data: 
            dfs.append(pd.DataFrame(resp.data))
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame(columns=["codigo_produto", "loja", "disponivel"])

@st.cache_data(ttl=300, show_spinner=False)
def carregar_produtos(setor: str, somente_ativos: bool = False) -> pd.DataFrame:
    # Catálogo do setor (muda pouco durante o dia). Cacheado por (setor, somente_ativos).
    supabase = obter_supabase()
    q = supabase.table("produtos").select("codigo, codigo_erp, codigo_iceasa, descricao, fornecedor, nome_personalizado").eq("setor", setor)
    if somente_ativos:
        q = q.eq("ativo", True)
    resp = q.execute()
    return pd.DataFrame(resp.data)

@st.cache_data(ttl=300, show_spinner=False)
def carregar_medias(num_loja: int) -> pd.DataFrame:
    # Médias 90d da loja. Só mudam quando o admin puxa do ERP (que limpa o cache).
    supabase = obter_supabase()
    resp = supabase.table("medias_90d").select("codigo_produto, media_dia").eq("loja", num_loja).execute()
    return pd.DataFrame(resp.data)

def sincronizar_medias_setor(setor: str, view_sql: str) -> tuple:
    # Puxa a média do período (view do ERP) e regrava medias_90d p/ os produtos do setor.
    # Reutilizada pelo botão "Puxar Médias do ERP" e pela atualização automática ao abrir a loja.
    # Retorna (sucesso: bool, qtd: int, msg: str).
    supabase = obter_supabase()
    resp_prod = supabase.table("produtos").select("codigo, codigo_erp").eq("setor", setor).execute()
    df_prod_map = pd.DataFrame(resp_prod.data)
    if df_prod_map.empty:
        return (False, 0, "Nenhum produto cadastrado neste setor.")
    df_prod_map = df_prod_map.rename(columns={'codigo': 'codigo_pk_interna'})
    if 'codigo_erp' not in df_prod_map.columns:
        df_prod_map['codigo_erp'] = df_prod_map['codigo_pk_interna']
    df_prod_map['codigo_erp'] = df_prod_map['codigo_erp'].fillna(df_prod_map['codigo_pk_interna']).astype(int)
    codigos_erp_setor = df_prod_map['codigo_erp'].unique().tolist()
    df_erp = conn_pg.query(f'SELECT * FROM "{view_sql}"', ttl=0)
    if df_erp.empty:
        return (False, 0, "View do ERP retornou vazia.")
    c_loja, c_cod_erp, c_med = df_erp.columns[0], df_erp.columns[1], df_erp.columns[2]
    df_erp_setor = df_erp[df_erp[c_cod_erp].isin(codigos_erp_setor)]
    if df_erp_setor.empty:
        return (False, 0, "View vazia para estes produtos.")
    df_merged = pd.merge(df_erp_setor, df_prod_map, left_on=c_cod_erp, right_on='codigo_erp', how='inner')
    codigos_pks = df_merged['codigo_pk_interna'].unique().tolist()
    for i in range(0, len(codigos_pks), 200):
        supabase.table("medias_90d").delete().in_("codigo_produto", codigos_pks[i:i+200]).execute()
    lista_insert = []
    for _, row in df_merged.iterrows():
        lista_insert.append({
            "loja": int(row[c_loja]),
            "codigo_produto": int(row['codigo_pk_interna']),
            "media_dia": float(row[c_med]) if pd.notna(row[c_med]) else 0.0
        })
    for i in range(0, len(lista_insert), 1000):
        supabase.table("medias_90d").insert(lista_insert[i:i+1000]).execute()
    return (True, len(lista_insert), "ok")

def carregar_extras(setor: str, data_str: str) -> pd.DataFrame:
    # Preço/observação do dia (preenchidos manualmente na Separação). SEM cache, pois
    # mudam quando o comprador digita e salva — precisa refletir na hora.
    supabase = obter_supabase()
    resp = supabase.table("separacao_extras").select("codigo_produto, preco, observacao").eq("setor", setor).eq("data_pedido", data_str).execute()
    return pd.DataFrame(resp.data)

def carregar_obs_loja(setor: str, num_loja: int, data_str: str) -> str:
    # Observação Geral que a própria loja digita (embalagem/padaria/confeitaria). SEM cache.
    supabase = obter_supabase()
    resp = supabase.table("observacoes_lojas").select("observacao").eq("setor", setor).eq("loja", num_loja).eq("data_pedido", data_str).execute()
    if resp.data:
        return (resp.data[0].get("observacao") or "")
    return ""

def carregar_obs_lojas_admin(setor: str, data_str: str) -> pd.DataFrame:
    # Todas as observações das lojas do dia, p/ o admin ver na Separação. SEM cache.
    supabase = obter_supabase()
    resp = supabase.table("observacoes_lojas").select("loja, observacao").eq("setor", setor).eq("data_pedido", data_str).execute()
    return pd.DataFrame(resp.data)

def salvar_obs_loja(setor: str, num_loja: int, data_str: str, texto: str, usuario: str):
    # Regrava a observação da loja (apaga e insere se houver texto). SEM cache.
    supabase = obter_supabase()
    supabase.table("observacoes_lojas").delete().eq("setor", setor).eq("loja", num_loja).eq("data_pedido", data_str).execute()
    txt = (texto or "").strip()
    if txt:
        supabase.table("observacoes_lojas").insert({
            "data_pedido": data_str, "setor": setor, "loja": num_loja,
            "observacao": txt, "usuario": usuario
        }).execute()

# 🚫 "Sem Pedido Hoje" — registro de que a loja declarou que NÃO fará pedido no dia.
# Guardado na tabela sem_pedido_hoje p/ a Separação poder marcar a loja em verde
# (senão, como o pedido é apagado, ela voltaria a aparecer como "Faltando").
def carregar_sem_pedido(setor: str, data_str: str) -> set:
    # Conjunto de lojas (int) que declararam "Sem Pedido Hoje" no setor/dia. SEM cache.
    # Defensivo: se a tabela ainda não existir, retorna vazio (não quebra a tela).
    try:
        supabase = obter_supabase()
        resp = supabase.table("sem_pedido_hoje").select("loja").eq("setor", setor).eq("data_pedido", data_str).execute()
        return {int(r["loja"]) for r in resp.data} if resp.data else set()
    except Exception:
        return set()

def registrar_sem_pedido(setor: str, num_loja: int, data_str: str, usuario: str):
    # Grava (regravando) a declaração de "sem pedido" da loja no dia.
    supabase = obter_supabase()
    try:
        supabase.table("sem_pedido_hoje").delete().eq("setor", setor).eq("loja", num_loja).eq("data_pedido", data_str).execute()
        supabase.table("sem_pedido_hoje").insert({
            "data_pedido": data_str, "setor": setor, "loja": num_loja, "usuario": usuario
        }).execute()
    except Exception as e:
        st.warning(f"⚠️ Não consegui marcar 'Sem Pedido' na Separação — crie a tabela `sem_pedido_hoje` no Supabase. ({e})")

def remover_sem_pedido(setor: str, num_loja: int, data_str: str):
    # Remove a marca de "sem pedido" (ex.: a loja mudou de ideia e lançou itens).
    try:
        supabase = obter_supabase()
        supabase.table("sem_pedido_hoje").delete().eq("setor", setor).eq("loja", num_loja).eq("data_pedido", data_str).execute()
    except Exception:
        pass

# ─────────────────────────────────────────────────────────────────────────────
# 🛡️ PROTEÇÃO CONTRA PERDA DE DIGITAÇÃO (avisa antes de fechar/recarregar a aba)
# ─────────────────────────────────────────────────────────────────────────────
def guardar_contra_saida(tem_alteracoes: bool):
    if tem_alteracoes:
        st.warning("⚠️ Há alterações **não salvas** nesta tela. Clique em **Salvar** antes de sair, recarregar ou trocar de tela.")
        st.components.v1.html(
            "<script>window.parent.onbeforeunload=function(e){e.preventDefault();e.returnValue='';return '';};</script>",
            height=0,
        )
    else:
        st.components.v1.html(
            "<script>window.parent.onbeforeunload=null;</script>",
            height=0,
        )

# ─────────────────────────────────────────────────────────────────────────────
# 🪟 MODAIS DE CONFIRMAÇÃO (janelas centralizadas, aparecem sobre a tela)
# ─────────────────────────────────────────────────────────────────────────────
@st.dialog("🚨 Confirmação Necessária")
def modal_limpar_pedidos(setor: str):
    # Limpa os pedidos do setor no dia (todas as lojas), além de preços/observações.
    st.markdown(f"Tem certeza que deseja **limpar todos os pedidos** de **{setor}** de hoje?")
    st.markdown("⚠️ *Esta ação apaga os pedidos, preços/observações e observações das lojas do dia — de todas as lojas — e não pode ser desfeita.*")
    st.write("<br>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("❌ Não, cancelar", use_container_width=True, key=f"mlp_nao_{setor}"):
            st.rerun()
    with c2:
        if st.button("✔️ Sim, limpar tudo", type="primary", use_container_width=True, key=f"mlp_sim_{setor}"):
            supabase = obter_supabase()
            with st.spinner("Limpando..."):
                supabase.table("pedidos").delete().eq("setor", setor).eq("data_pedido", str(data_brasilia())).execute()
                supabase.table("separacao_extras").delete().eq("setor", setor).eq("data_pedido", str(data_brasilia())).execute()
                supabase.table("observacoes_lojas").delete().eq("setor", setor).eq("data_pedido", str(data_brasilia())).execute()
                try:
                    supabase.table("sem_pedido_hoje").delete().eq("setor", setor).eq("data_pedido", str(data_brasilia())).execute()
                except Exception:
                    pass
            # reseta o editor p/ não ficar 'None' preso em células que foram apagadas
            st.session_state.pop("editor_separacao", None)
            st.rerun()

@st.dialog("🚫 Confirmação — Sem Pedido Hoje")
def modal_sem_pedido(setor: str, num_loja: int, loja_selecionada: str):
    # Zera o pedido da loja no dia e avisa o supervisor no Telegram.
    st.markdown(f"Confirma que a **{loja_selecionada}** **NÃO** fará pedido hoje?")
    st.markdown("⚠️ *Isso apaga o que estiver lançado e avisa o supervisor no Telegram.*")
    st.write("<br>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("❌ Não, cancelar", use_container_width=True, key=f"msp_nao_{setor}_{num_loja}"):
            st.rerun()
    with c2:
        if st.button("✔️ Sim, sem pedido hoje", type="primary", use_container_width=True, key=f"msp_sim_{setor}_{num_loja}"):
            supabase = obter_supabase()
            with st.spinner("Registrando 'Sem Pedido Hoje'..."):
                supabase.table("pedidos").delete().eq("setor", setor).eq("loja", num_loja).eq("data_pedido", str(data_brasilia())).execute()
                # registra a declaração p/ a loja ficar VERDE ("Sem Pedido") na Separação
                registrar_sem_pedido(setor, num_loja, str(data_brasilia()), st.session_state.get('usuario_logado', loja_selecionada))
                msg_aviso = (
                    f"🚨 AVISO - {setor}\n"
                    f"A {loja_selecionada} informou que NÃO fará pedido hoje "
                    f"({data_hora_brasilia()})."
                )
                enviado = notificar_telegram(msg_aviso)
            st.session_state[f"sem_pedido_msg_{setor}_{num_loja}"] = "ok" if enviado else "parcial"
            # limpa o editor (todas as variações de filtro) para a tela voltar zerada
            for _k in [k for k in list(st.session_state.keys()) if str(k).startswith(f"grid_loja_{num_loja}")]:
                st.session_state.pop(_k, None)
            st.rerun()

@st.dialog("✅ Pedido Salvo com Sucesso")
def modal_pedido_salvo(loja_selecionada: str):
    # Janela de confirmação exibida quando a loja salva o pedido e a gravação dá certo.
    # É disparada após o rerun do botão "Salvar Pedido" (via flag no session_state).
    st.markdown(f"O pedido da **{loja_selecionada}** foi **salvo com sucesso**! ✅")
    st.markdown("Pode conferir, imprimir ou exportar quando quiser.")
    st.write("<br>", unsafe_allow_html=True)
    if st.button("👍 Fechar", type="primary", use_container_width=True, key=f"modal_salvo_fechar_{loja_selecionada}"):
        st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# 🧠 FUNÇÃO DIRETORA DO MÓDULO UNIFICADO
# ─────────────────────────────────────────────────────────────────────────────
def iniciar_tela(setor: str):
    supabase = obter_supabase()
    usuario_atual = st.session_state.get('usuario_logado', 'Loja 01')
    acesso_total = (usuario_atual == "Administrador")

    st.markdown("""
        <style>
        div[data-testid="stComponentStack"] { width: 100% !important; }
        div[data-testid="stTable"] td { text-align: center !important; }
        
        [data-testid="stSidebar"] button[data-testid="stBaseButton-primary"],
        [data-testid="stSidebar"] button[kind="primary"] {
            background-color: #ff4b4b !important; 
            color: white !important; 
            border-color: #ff4b4b !important;
        }
        
        @media screen {
            .print-only { display: none !important; } 
            /* 📐 Usa a largura cheia da tela — tabelas largas (Separação/Catálogo) cabem sem rolar */
            [data-testid="stMainBlockContainer"],
            [data-testid="stAppViewBlockContainer"],
            .block-container {
                max-width: 100% !important;
                padding-left: 2rem !important;
                padding-right: 2rem !important;
            }
        }
        
        /* 🔥 CSS OTMIZADO PARA PREENCHIMENTO 100% NA IMPRESSÃO 🔥 */
        @media print {
            @page {
                margin: 10mm;
            }

            /* Oculta tudo que não é o relatório e zera seu espaço */
            section[data-testid="stSidebar"], 
            div[data-testid="stSidebarNav"],
            header[data-testid="stHeader"], 
            footer, 
            [data-testid="stToolbar"],
            button, .stButton, 
            [data-testid="stMetric"], 
            [data-testid="stRadio"], 
            [data-testid="stSelectbox"], 
            [data-testid="stTextInput"], 
            [data-testid="stAlert"],
            [data-testid="stHorizontalBlock"], /* Esconde as colunas de filtros/botões de cima */
            h1, h2, h5, h6, 
            [data-testid="stDataEditor"], 
            [data-testid="stDataFrame"],
            .no-print {
                display: none !important;
                width: 0 !important;
                height: 0 !important;
                min-width: 0 !important;
                margin: 0 !important;
                padding: 0 !important;
            }

            /* 🔥 CORREÇÃO 1 — ELIMINA O ESPAÇO EM BRANCO NO TOPO 🔥 */
            /* Esconde QUALQUER bloco de elemento do Streamlit que NÃO contenha a área
               de impressão (.print-only). Assim o relatório "sobe" sozinho para o topo
               da folha, independente de quantos data_editors/métricas/filtros existam
               acima dele. Continua paginando normalmente (não usa position absolute). */
            [data-testid="stMain"] [data-testid="stElementContainer"]:not(:has(.print-only)),
            [data-testid="stMain"] .element-container:not(:has(.print-only)) {
                display: none !important;
            }
            
            /* Remove os preenchimentos/bordas invisíveis dos st.container */
            [data-testid="stVerticalBlockBorderWrapper"] {
                border: none !important;
                padding: 0 !important;
                margin: 0 !important;
            }
            div[data-testid="stVerticalBlock"], .element-container {
                gap: 0 !important;
                margin-bottom: 0 !important;
                padding-bottom: 0 !important;
            }

            /* FORÇA O PREENCHIMENTO DO ESPAÇO DA ESQUERDA PARA 0 */
            /* DEPOIS */
            [data-testid="stMain"], .main,
            html, body, .stApp, 
            [data-testid="stAppViewContainer"],
            [data-testid="stAppViewBlockContainer"],
            .block-container {
                background-color: white !important;
                position: static !important;
                margin: 0 !important;
                padding: 0 !important;
                width: 100% !important;
                max-width: 100% !important;
                overflow: visible !important;
            }

            /* Puxa agressivamente o conteúdo pro topo */
            [data-testid="stMain"] > div:first-child,
            [data-testid="stAppViewBlockContainer"] > div:first-child {
                margin-top: -15px !important;
                padding-top: 0 !important;
            }
            
            .print-only {
                display: block !important;
                width: 100% !important;
            }
            
            .print-only h3 {
                font-size: 12pt !important;
                margin: 0 0 10px 0 !important;
                color: black !important;
            }

            /* Linha de data/hora (padrão brasileiro) impressa dentro do relatório */
            .print-datetime {
                font-size: 8.5pt !important;
                color: #333 !important;
                margin: -6px 0 8px 0 !important;
                text-align: left !important;
            }
            
            /* Observação Geral da Loja abaixo do pedido impresso */
            .print-obs-loja {
                margin-top: 10px !important;
                padding: 6px 8px !important;
                border: 1px solid #000 !important;
                font-size: 9pt !important;
                line-height: 1.3 !important;
                page-break-inside: avoid !important;
            }
            
            /* Título do fornecedor grudado na tabela como um bloco contínuo */
            .print-only h4.supplier-header {
                font-size: 9.5pt !important;
                margin: 0 !important;
                padding: 6px 8px !important;
                background-color: #e0e0e0 !important;
                border: 1px solid black !important;
                border-bottom: none !important;
                color: black !important;
                text-align: left !important;
            }
            
            /* Tabela ultra-comprimida */
            table.print-table {
                width: 100% !important;
                border-collapse: collapse;
                font-family: Arial, sans-serif;
                font-size: 8.5pt !important;
                margin-top: 0 !important;
                margin-bottom: 20px !important; /* 🔥 Aumentado o espaço entre fornecedores para não sobrepor 🔥 */
                page-break-inside: auto;
            }
            
            table.print-table tr {
                page-break-inside: avoid;
                page-break-after: auto;
            }
            
            table.print-table th, table.print-table td {
                border: 1px solid black !important;
                padding: 4px 6px !important; 
                color: black !important;
                line-height: 1.1 !important; 
                text-align: center;
            }
            
            table.print-table th {
                background-color: #f0f2f6 !important;
                font-weight: bold;
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }
            
            table.print-table td:nth-child(2), table.print-table td:nth-child(3), table.print-table td:nth-child(4) {
                text-align: left;
            }

            /* 🔥 CORREÇÃO 2 — VISÃO FORNECEDORES: fonte/padding menores p/ nomes não quebrarem 🔥 */
            /* (ex.: "Alface Hidrop Fukunaga Crespa Un" deixava de quebrar p/ a linha de baixo) */
            .print-fornecedores table.print-table {
                font-size: 7pt !important;
            }
            .print-fornecedores table.print-table th,
            .print-fornecedores table.print-table td {
                padding: 2px 4px !important;
                line-height: 1.05 !important;
            }
            /* As colunas de Loja/Total ficam estreitas → sobra largura p/ a coluna Produto */
            .print-fornecedores table.print-table th:nth-child(n+3),
            .print-fornecedores table.print-table td:nth-child(n+3) {
                width: 6% !important;
            }
            /* 🔥 AJUSTE — centraliza os pedidos digitados (Lojas/Total). 🔥
               Sobrescreve a regra global que jogava as colunas 2/3/4 p/ a esquerda.
               Só a coluna "Produto" (2ª) permanece alinhada à esquerda. */
            .print-fornecedores table.print-table th,
            .print-fornecedores table.print-table td {
                text-align: center !important;
            }
            .print-fornecedores table.print-table th:nth-child(2),
            .print-fornecedores table.print-table td:nth-child(2) {
                text-align: left !important;
            }

            /* 🔥 CORREÇÃO 3 — VISÃO DAS LOJAS: larguras fixas p/ não cortar "Qtde Pedida" 🔥 */
            /* table-layout: fixed faz o navegador respeitar as larguras abaixo (soma = 100%)
               e quebrar a Descrição em 2 linhas em vez de empurrar a última coluna p/ fora. */
            .print-lojas table.print-table {
                table-layout: fixed !important;
                width: 100% !important;
                font-size: 8pt !important;
            }
            .print-lojas table.print-table th:nth-child(1),
            .print-lojas table.print-table td:nth-child(1) { width: 7%; }   /* Cód. ERP    */
            .print-lojas table.print-table th:nth-child(2),
            .print-lojas table.print-table td:nth-child(2) { width: 13%; }  /* Fornecedor  */
            .print-lojas table.print-table th:nth-child(3),
            .print-lojas table.print-table td:nth-child(3) { width: 51%; }  /* Descrição   */
            .print-lojas table.print-table th:nth-child(4),
            .print-lojas table.print-table td:nth-child(4) { width: 9%; }   /* Média (90d) */
            .print-lojas table.print-table th:nth-child(5),
            .print-lojas table.print-table td:nth-child(5) { width: 9%; }   /* Estoque ERP */
            .print-lojas table.print-table th:nth-child(6),
            .print-lojas table.print-table td:nth-child(6) { width: 11%; }  /* Pedido      */
            .print-lojas table.print-table td {
                word-break: break-word;
                overflow-wrap: anywhere;
            }
            /* Números (Média/Estoque/Qtde) centralizados; Fornecedor/Descrição à esquerda */
            .print-lojas table.print-table th:nth-child(4),
            .print-lojas table.print-table td:nth-child(4),
            .print-lojas table.print-table th:nth-child(5),
            .print-lojas table.print-table td:nth-child(5),
            .print-lojas table.print-table th:nth-child(6),
            .print-lojas table.print-table td:nth-child(6) {
                text-align: center !important;
            }
        }
        </style>
    """, unsafe_allow_html=True)

    with st.sidebar:
        st.markdown(f"### Parâmetros: {setor}")
        if acesso_total:
            if setor_usa_iceasa(setor):
                # FLV Normal/Ofertas entram direto na Separação e Fechamento
                opcoes_nav = [
                    "Separação e Fechamento",
                    "Pedido por Fornecedor",
                    "Visão Fornecedores (Resumo)",
                    "Visão das Lojas",
                    "Catálogo de Produtos",
                ]
            else:
                opcoes_nav = [
                    "Visão Fornecedores (Resumo)",
                    "Separação e Fechamento",
                    "Visão das Lojas",
                    "Catálogo de Produtos",
                ]
            perfil_navegacao = st.radio("📍 Navegação Interna:", opcoes_nav)
        else:
            perfil_navegacao = "Visão das Lojas"

    # 🔥 INJEÇÃO PARA MODO PAISAGEM (LANDSCAPE) APENAS NA SEPARAÇÃO E FECHAMENTO 🔥
    if perfil_navegacao == "Separação e Fechamento":
        st.markdown("""
            <style>
            @media print {
                @page { size: landscape; margin: 10mm; }
            }
            </style>
        """, unsafe_allow_html=True)

    if acesso_total:
        with st.sidebar:
            st.markdown("---")
            # 🔄 Sincronizar Dados: recarrega os dados mais recentes (pedidos que as lojas
            # acabaram de digitar, catálogo e médias). Limpa os caches e refaz a tela.
            if st.button("🔄 Sincronizar Dados", use_container_width=True,
                         help="Atualiza a tela com os dados mais recentes: pedidos digitados pelas lojas, catálogo e médias."):
                st.cache_data.clear()
                st.rerun()

            st.markdown("---")
            st.markdown("🔄 **Atualizar Médias (90d)**")
            
            opcoes_views = list(VIEWS_MEDIA.keys())
            # pré-seleciona o filtro padrão do setor (o admin ainda pode trocar)
            filtro_padrao = filtro_media_padrao(setor)
            idx_padrao = opcoes_views.index(filtro_padrao) if filtro_padrao in opcoes_views else 0
            view_escolhida = st.selectbox("Selecione o período base:", opcoes_views, index=idx_padrao, label_visibility="collapsed")
            
            if st.button("📥 Puxar Médias do ERP", type="secondary", use_container_width=True):
                view_sql = VIEWS_MEDIA[view_escolhida]
                with st.spinner(f"Sincronizando {view_sql}..."):
                    try:
                        ok, qtd, msg = sincronizar_medias_setor(setor, view_sql)
                        if ok:
                            st.success(f"Médias ({view_escolhida}) sincronizadas!")
                            st.cache_data.clear()  # médias mudaram → recarrega na hora
                            time.sleep(1.5)
                            st.rerun()
                        else:
                            st.warning(msg)
                    except Exception as e: 
                        st.error(f"Erro na sincronização: {e}")

            st.markdown("---")
            if st.button("🗑️ Limpar Pedidos", type="primary", use_container_width=True):
                modal_limpar_pedidos(setor)

    # ─────────────────────────────────────────────────────────────────────────
    # ROTA 1 — SEPARAÇÃO E FECHAMENTO
    # ─────────────────────────────────────────────────────────────────────────
    if perfil_navegacao == "Separação e Fechamento":
        st.markdown(f"<div class='no-print'><h2>📊 Separação e Fechamento — {setor}</h2></div>", unsafe_allow_html=True)
        
        df_prod = carregar_produtos(setor).copy()
        resp_ped = supabase.table("pedidos").select("codigo_produto, loja, quantidade").eq("setor", setor).eq("data_pedido", str(data_brasilia())).execute()
        
        df_ped = pd.DataFrame(resp_ped.data)
        
        if df_prod.empty: 
            st.warning("Nenhum produto cadastrado para este setor.")
            return
            
        if 'codigo_erp' not in df_prod.columns: 
            df_prod['codigo_erp'] = df_prod['codigo']
            
        df_prod['codigo_erp'] = df_prod['codigo_erp'].fillna(df_prod['codigo']).astype(int)

        codigos_setor = df_prod['codigo'].tolist()
        df_perm_all = buscar_permissoes_setor(supabase, codigos_setor)
        df_prod['descricao'] = df_prod['nome_personalizado'].apply(lambda x: str(x).strip() if pd.notna(x) and str(x).strip() != "" else None).fillna(df_prod['descricao'])

        # Lojas que declararam "Sem Pedido Hoje" (Açougue Adriano / Especiais) → ficam verdes
        lojas_sem_pedido = carregar_sem_pedido(setor, str(data_brasilia())) if setor_usa_sem_pedido(setor) else set()
        exibir_status_digitacao_lojas(df_ped, lojas_sem_pedido)
        
        if not df_ped.empty:
            df_pivot = df_ped.pivot_table(index='codigo_produto', columns='loja', values='quantidade', aggfunc='sum').reset_index()
            for n in range(1, 9):
                if n in df_pivot.columns: 
                    df_pivot = df_pivot.rename(columns={n: f"Loja {n:02d}"})
        else: 
            df_pivot = pd.DataFrame(columns=['codigo_produto'])

        df_consolidado = pd.merge(df_prod, df_pivot, left_on='codigo', right_on='codigo_produto', how='left')
        
        if not df_perm_all.empty:
            df_perm_all['loja_nome'] = df_perm_all['loja'].apply(lambda x: f"Loja {int(x):02d}_perm")
            df_perm_pivot = df_perm_all.pivot_table(index='codigo_produto', columns='loja_nome', values='disponivel', aggfunc='last').reset_index()
        else: 
            df_perm_pivot = pd.DataFrame(columns=['codigo_produto'])

        df_consolidado = pd.merge(df_consolidado, df_perm_pivot, left_on='codigo', right_on='codigo_produto', how='left')
        
        perm_cols = [f"Loja {n:02d}_perm" for n in range(1, 9)]
        for col in perm_cols:
            if col not in df_consolidado.columns: 
                df_consolidado[col] = True
                
        df_consolidado[perm_cols] = df_consolidado[perm_cols].fillna(True)
        mask_active = df_consolidado[perm_cols].any(axis=1)
        df_consolidado = df_consolidado[mask_active]

        for loja in LOJAS_NOMES:
            if loja not in df_consolidado.columns: 
                df_consolidado[loja] = 0.0
            df_consolidado[loja] = df_consolidado[loja].fillna(0).astype(int)

        df_consolidado["TOTAL_NUM"] = df_consolidado[LOJAS_NOMES].sum(axis=1)

        st.markdown("<div class='no-print'><br></div>", unsafe_allow_html=True)
        col_filtro, col_metric = st.columns([3, 1])
        with col_filtro:
            # Agrupa "BIG FRANGO - *" sob o rótulo "BIG FRANGO" (igual à Visão das Lojas).
            # A coluna Fornecedor segue mostrando o tipo específico de cada produto.
            df_consolidado["_grupo_forn"] = df_consolidado['fornecedor'].apply(grupo_fornecedor)
            opcoes_forn = ["Todos"] + sorted([g for g in df_consolidado["_grupo_forn"].dropna().unique() if str(g).strip() != ""])
            filtro_selecionado = st.radio("🔍 Filtrar Exibição por Setor:", options=opcoes_forn, horizontal=True)

        if filtro_selecionado != "Todos": 
            df_consolidado = df_consolidado[df_consolidado['_grupo_forn'] == filtro_selecionado]

        st.metric(label="📦 Itens c/ pedido", value=f"{df_consolidado[df_consolidado['TOTAL_NUM'] > 0].shape[0]} produtos")

        df_consolidado["TOTAL GERAL"] = df_consolidado["TOTAL_NUM"].replace({0: ""})
        for loja in LOJAS_NOMES:
            df_consolidado[loja] = df_consolidado[loja].replace({0: ""}).astype(str).replace({"0": "", "0.0": "", "nan": ""})
            perm_col = f"{loja}_perm"
            if perm_col in df_consolidado.columns: 
                df_consolidado.loc[df_consolidado[perm_col] != True, loja] = "-"

        df_consolidado = df_consolidado.rename(columns={'codigo_erp': 'Cód. ERP', 'codigo_iceasa': 'Cód. Iceasa', 'descricao': 'Descrição', 'fornecedor': 'Fornecedor'})

        usa_iceasa = setor_usa_iceasa(setor)
        usa_erp = setor_usa_erp(setor)
        texto_setor = setor_pedido_texto(setor)
        eh_mp = setor_eh_materia_prima(setor)
        if usa_iceasa:
            cols_cod = ["Cód. ERP", "Cód. Iceasa"]
        elif usa_erp:
            cols_cod = ["Cód. ERP"]
        else:
            cols_cod = []   # Peças Açougue - Manoel: sem código ERP
        if usa_iceasa and "Cód. Iceasa" not in df_consolidado.columns:
            df_consolidado["Cód. Iceasa"] = None

        # 💲 Preço/Observação do dia: FLV tem Preço+Observação; Matéria Prima tem só Observação.
        # Tudo guardado em separacao_extras (sem tabela nova) e mapeado por 'codigo'.
        cols_extras = []
        if usa_iceasa or eh_mp:
            df_extras = carregar_extras(setor, str(data_brasilia()))
            mapa_obs = dict(zip(df_extras["codigo_produto"], df_extras["observacao"])) if not df_extras.empty else {}
            df_consolidado["Observação"] = df_consolidado["codigo"].map(mapa_obs).fillna("").astype(str).replace({"None": "", "nan": "", "<NA>": ""})
            if usa_iceasa:
                mapa_preco = dict(zip(df_extras["codigo_produto"], df_extras["preco"])) if not df_extras.empty else {}
                df_consolidado["R$ Preço"] = df_consolidado["codigo"].map(mapa_preco).apply(preco_para_celula)
                cols_extras = ["R$ Preço", "Observação"]
            else:
                cols_extras = ["Observação"]

        # Matéria Prima/Embalagem/Padaria: sem coluna Total (pedem em texto)
        # Matéria Prima/Embalagem/Padaria: sem coluna Total em tudo (tela, Excel e impressão)
        cols_total = [] if texto_setor else ["TOTAL GERAL"]
        df_exibicao = df_consolidado[["Fornecedor", "codigo"] + cols_cod + ["Descrição"] + LOJAS_NOMES + cols_total + cols_extras].sort_values(by=['Fornecedor', 'Descrição'])

        col_cfg = {
            "codigo": None, 
            "Cód. ERP": st.column_config.NumberColumn("Cód. ERP", disabled=True, format="%d", width=70), 
            "Cód. Iceasa": st.column_config.NumberColumn("Cód. Iceasa", disabled=True, format="%d", width=78), 
            "Fornecedor": st.column_config.TextColumn(disabled=True, width=180), 
            "Descrição": st.column_config.TextColumn(disabled=True, width=175), 
            "TOTAL GERAL": st.column_config.TextColumn("TOTAL", disabled=True, width=56)
        }
        if usa_iceasa:
            col_cfg["R$ Preço"] = st.column_config.TextColumn("R$ Preço", width=88, help="Preço do dia (ex.: 12,50). Deixe vazio se não houver.")
            col_cfg["Observação"] = st.column_config.TextColumn("Observação", width=145)
        elif eh_mp:
            col_cfg["Observação"] = st.column_config.TextColumn("Observação:", width=240, help="Observação do item — sai também na exportação Excel.")
        for loja in LOJAS_NOMES: 
            col_cfg[loja] = st.column_config.TextColumn(loja, width=72, disabled=False)

        # 🧹 Antes de desenhar o editor: troca por vazio qualquer célula de Preço/Observação
        # que o usuário apagou (o data_editor guarda 'None' no estado e mostra "None" na tela).
        # Mexer no estado ANTES de instanciar o widget é permitido e faz a célula voltar a vazio.
        if usa_iceasa or eh_mp:
            _est = st.session_state.get("editor_separacao")
            if isinstance(_est, dict) and isinstance(_est.get("edited_rows"), dict):
                for _ch in _est["edited_rows"].values():
                    for _c in ("R$ Preço", "Observação"):
                        if _c in _ch and _ch[_c] is None:
                            _ch[_c] = ""

        df_editado = st.data_editor(df_exibicao, hide_index=True, use_container_width=True, height=500, column_config=col_cfg, key="editor_separacao")

        # 🛡️ Alterações não salvas: compara lojas + (preço/obs no FLV)
        cols_guarda = LOJAS_NOMES + cols_extras
        orig_sep = df_exibicao[cols_guarda].fillna("").astype(str).values.tolist()
        edit_sep = df_editado[cols_guarda].fillna("").astype(str).values.tolist()
        guardar_contra_saida(orig_sep != edit_sep)
        
        # 🔥 Impressão: troca "-" por vazio; FLV imprime o Iceasa (ERP sai). Preço/Obs ficam fora da impressão por ora.
        df_print_sep = df_exibicao.drop(columns=['codigo', 'R$ Preço', 'Observação'], errors='ignore').fillna('').replace("-", "")
        if usa_iceasa:
            df_print_sep['Cód. Iceasa'] = df_print_sep['Cód. Iceasa'].apply(iceasa_para_impressao)
            df_print_sep = df_print_sep.drop(columns=['Cód. ERP'], errors='ignore')
        html_table = df_print_sep.to_html(index=False, classes="print-table")
        st.markdown(f'<div class="print-only"><h3>📊 Separação e Fechamento — {setor} ({filtro_selecionado})</h3><div class="print-datetime">Emitido em {data_hora_brasilia()}</div>{html_table}</div>', unsafe_allow_html=True)

        c_salvar, c_excel, c_print = st.columns([2, 2, 1])
        with c_salvar: 
            btn_salvar = st.button("💾 Salvar Ajustes Administrativos", type="primary", use_container_width=True)
        with c_excel:
            df_export = df_editado.drop(columns=['codigo'], errors='ignore')
            # 📦 FLV Normal e FLV Ofertas usam o modelo "PEDIDO BOX"; demais setores seguem o padrão
            if str(setor).strip().lower() in ("flv normal", "flv ofertas"):
                excel_bytes = gerar_excel_box(df_export)
                nome_arq = "molicenter.xlsx"
            else:
                excel_bytes = gerar_excel_download(df_export, f"Fechamento {setor}", com_obs=setor_eh_materia_prima(setor))
                nome_arq = f"Separacao_Fechamento_{setor}.xlsx"
            st.download_button("📊 Exportar Excel", data=excel_bytes, file_name=nome_arq, use_container_width=True)
        with c_print: 
            injetar_botao_impressao()
            
        if btn_salvar:
            with st.spinner("Atualizando registros..."):
                cods = df_editado["codigo"].tolist()
                if cods:
                    for loja_nome in LOJAS_NOMES:
                        n_loja = int(loja_nome.split()[-1])
                        supabase.table("pedidos").delete().eq("setor", setor).eq("loja", n_loja).eq("data_pedido", str(data_brasilia())).in_("codigo_produto", cods).execute()
                        
                        lista_ins = []
                        for _, r in df_editado.iterrows():
                            q = converter_para_int_seguro(r[loja_nome])
                            if q > 0: 
                                lista_ins.append({"data_pedido": str(data_brasilia()), "setor": setor, "loja": n_loja, "codigo_produto": int(r["codigo"]), "quantidade": q, "usuario": usuario_atual})
                        
                        if lista_ins: 
                            supabase.table("pedidos").insert(lista_ins).execute()

                    # 💲 Preço/Observação do dia: FLV grava Preço+Obs; Matéria Prima grava só Obs.
                    if usa_iceasa or eh_mp:
                        for i in range(0, len(cods), 200):
                            supabase.table("separacao_extras").delete().eq("setor", setor).eq("data_pedido", str(data_brasilia())).in_("codigo_produto", cods[i:i+200]).execute()
                        lista_extras = []
                        for _, r in df_editado.iterrows():
                            preco_val = celula_para_preco(r.get("R$ Preço")) if usa_iceasa else None
                            obs_raw = r.get("Observação")
                            obs_val = str(obs_raw).strip() if pd.notna(obs_raw) and str(obs_raw).strip() != "" else None
                            if preco_val is not None or obs_val is not None:
                                lista_extras.append({"codigo_produto": int(r["codigo"]), "data_pedido": str(data_brasilia()), "setor": setor, "preco": preco_val, "observacao": obs_val})
                        for i in range(0, len(lista_extras), 1000):
                            supabase.table("separacao_extras").insert(lista_extras[i:i+1000]).execute()
            st.success("Alterações consolidadas!")
            # limpa o estado do editor para o guarda de "não salvo" desligar
            st.session_state.pop("editor_separacao", None)
            st.rerun()

        # 📝 Observações das Lojas (embalagem/padaria/confeitaria): o admin vê o recado de cada loja
        if setor_usa_obs_loja(setor) and acesso_total:
            df_obs = carregar_obs_lojas_admin(setor, str(data_brasilia()))
            st.markdown("<div class='no-print'><br></div>", unsafe_allow_html=True)
            st.markdown(f"<div class='no-print'><h3>📝 Observações das Lojas — {data_hora_brasilia()}</h3></div>", unsafe_allow_html=True)
            if df_obs is not None and not df_obs.empty:
                df_obs = df_obs.copy()
                df_obs["Loja"] = df_obs["loja"].apply(lambda n: f"Loja {int(n):02d}")
                df_obs = df_obs.rename(columns={"observacao": "Observação"})[["Loja", "Observação"]].sort_values("Loja")
                st.dataframe(
                    df_obs, hide_index=True, use_container_width=True,
                    column_config={
                        "Loja": st.column_config.TextColumn(width=90),
                        "Observação": st.column_config.TextColumn(width=700),
                    },
                )
                html_obs = df_obs.to_html(index=False, classes="print-table")
                st.markdown(f'<div class="print-only"><h3>📝 Observações das Lojas</h3>{html_obs}</div>', unsafe_allow_html=True)
            else:
                st.info("Nenhuma loja registrou observação para hoje.")

    # ─────────────────────────────────────────────────────────────────────────
    # ROTA 2 — VISÃO DAS LOJAS
    # ─────────────────────────────────────────────────────────────────────────
    elif perfil_navegacao == "Visão das Lojas":
        loja_selecionada = st.selectbox("👁️ Visualizar como:", LOJAS_NOMES) if acesso_total else usuario_atual
        num_loja = int(loja_selecionada.split()[-1])
        usa_media = setor_usa_media(setor)
        usa_sem_pedido = setor_usa_sem_pedido(setor)

        # Mensagem do "Sem Pedido Hoje" — guardada no estado e exibida após o rerun
        _msg_sp = st.session_state.pop(f"sem_pedido_msg_{setor}_{num_loja}", None)
        if _msg_sp == "ok":
            st.success("✅ Supervisor avisado e pedido zerado! Nenhum pedido será feito hoje.")
        elif _msg_sp == "parcial":
            st.warning("⚠️ O pedido foi zerado, mas falhou ao enviar o aviso no Telegram.")

        # ✅ Confirmação de salvamento — abre o modal após o rerun do "Salvar Pedido"
        if st.session_state.pop(f"pedido_salvo_ok_{setor}_{num_loja}", None):
            modal_pedido_salvo(loja_selecionada)

        st.markdown(f"<div class='no-print'><h2>🥬 Lançamento de Pedidos — {loja_selecionada}</h2></div>", unsafe_allow_html=True)
        
        df_prod = carregar_produtos(setor, somente_ativos=True).copy()

        if df_prod.empty: 
            st.warning("Nenhum produto cadastrado para este setor.")
            return

        if 'codigo_erp' not in df_prod.columns: 
            df_prod['codigo_erp'] = df_prod['codigo']
            
        df_prod['codigo_erp'] = df_prod['codigo_erp'].fillna(df_prod['codigo']).astype(int)

        codigos_setor = df_prod['codigo'].tolist()
        df_perm = buscar_permissoes_setor(supabase, codigos_setor, num_loja)
        
        # 🔄 Médias automáticas: ao abrir a loja já traz a média do filtro padrão do setor
        # (1x por sessão por setor+filtro; o admin ainda pode trocar/puxar manualmente na lateral).
        if acesso_total and usa_media:
            filtro_pad = filtro_media_padrao(setor)
            flag_med = f"med_auto_{setor}_{filtro_pad}"
            if not st.session_state.get(flag_med):
                try:
                    with st.spinner(f"Atualizando médias ({filtro_pad})..."):
                        ok_auto, _, _ = sincronizar_medias_setor(setor, VIEWS_MEDIA[filtro_pad])
                    if ok_auto:
                        carregar_medias.clear()
                except Exception:
                    pass  # ERP indisponível → segue com as médias já armazenadas
                st.session_state[flag_med] = True

        df_med = carregar_medias(num_loja).copy()
        resp_exis = supabase.table("pedidos").select("codigo_produto, quantidade").eq("setor", setor).eq("loja", num_loja).eq("data_pedido", str(data_brasilia())).execute()

        df_exis = pd.DataFrame(resp_exis.data)

        # 📝 Observação Geral da Loja — só embalagem/padaria/confeitaria
        usa_obs = setor_usa_obs_loja(setor)
        obs_atual = carregar_obs_loja(setor, num_loja, str(data_brasilia())) if usa_obs else ""

        df_prod['descricao'] = df_prod['nome_personalizado'].apply(lambda x: str(x).strip() if pd.notna(x) and str(x).strip() != "" else None).fillna(df_prod['descricao'])

        df_loja = pd.merge(df_prod, df_perm, left_on='codigo', right_on='codigo_produto', how='left')
        if 'disponivel' not in df_loja.columns: 
            df_loja['disponivel'] = True
        else: 
            df_loja['disponivel'] = df_loja['disponivel'].fillna(True)
            
        df_loja = df_loja[df_loja['disponivel'] == True]

        if df_loja.empty: 
            st.warning("Nenhum produto liberado para esta loja neste setor.")
            return

        df_loja = pd.merge(df_loja, df_med, on='codigo_produto', how='left')
        df_loja['media_dia'] = df_loja['media_dia'].fillna(0.0)

        if not df_exis.empty:
            df_loja = pd.merge(df_loja, df_exis, on='codigo_produto', how='left')
            df_loja['quantidade'] = df_loja['quantidade'].fillna(0).astype(int)
            itens_digitados = df_exis[df_exis['quantidade'] > 0].shape[0]
        else:
            df_loja['quantidade'] = 0
            itens_digitados = 0

        st.metric(label="📝 Seus Itens Preenchidos", value=f"{itens_digitados} produtos")

        usa_erp = setor_usa_erp(setor)
        if usa_erp:
            codigos_busca_erp = df_loja["codigo_erp"].dropna().astype(int).unique().tolist()
            df_estoque = buscar_estoque_erp(loja_selecionada, codigos_busca_erp, setor)
            df_loja = pd.merge(df_loja, df_estoque, left_on='codigo_erp', right_on='Código', how='left')
            df_loja["Estoque"] = df_loja["Estoque"].fillna(0).astype(int)
        else:
            # Peças Açougue - Manoel: não há ERP para puxar estoque
            df_loja["Estoque"] = 0
        
        df_loja['quantidade'] = df_loja['quantidade'].replace({0: ""}).astype(str).replace({"0": "", "0.0": "", "nan": ""})

        df_final_grid = pd.DataFrame({
            'codigo': df_loja['codigo'], 
            'Cód. ERP': df_loja['codigo_erp'],
            'Fornecedor': df_loja['fornecedor'], 
            'Descrição': df_loja['descricao'],
            'Média (90d)': df_loja['media_dia'], 
            'Estoque ERP': df_loja['Estoque'],
            'Qtde Pedida': df_loja['quantidade']
        }).sort_values(by=['Fornecedor', 'Descrição'])

        # Embalagem/Padaria/Confeitaria/Matéria Prima: sem a coluna Média na Visão das Lojas
        if not usa_media:
            df_final_grid = df_final_grid.drop(columns=['Média (90d)'], errors='ignore')

        # Peças Açougue - Manoel: sem Estoque ERP nem Cód. ERP (não há ERP para este setor)
        if not usa_erp:
            df_final_grid = df_final_grid.drop(columns=['Estoque ERP', 'Cód. ERP'], errors='ignore')

        # 🔍 Filtro por Fornecedor (igual ao da Separação). Os tipos "BIG FRANGO - *"
        # entram todos sob o mesmo rótulo "BIG FRANGO"; a coluna Fornecedor segue
        # mostrando o tipo específico. Só aparece quando há mais de um fornecedor.
        filtro_forn = "Todos"
        df_final_grid["_grupo_forn"] = df_final_grid["Fornecedor"].apply(grupo_fornecedor)
        grupos_forn = ["Todos"] + sorted(df_final_grid["_grupo_forn"].dropna().unique().tolist())
        if len(grupos_forn) > 2:
            filtro_forn = st.radio("🔍 Filtrar por Fornecedor:", options=grupos_forn, horizontal=True, key=f"filtro_forn_loja_{setor}_{num_loja}")
            if filtro_forn != "Todos":
                df_final_grid = df_final_grid[df_final_grid["_grupo_forn"] == filtro_forn]
        df_final_grid = df_final_grid.drop(columns=["_grupo_forn"], errors="ignore")

        texto_busca = st.text_input("🔍 Buscar Produto (por Código ou Nome):")
        if texto_busca:
            tb = texto_busca.lower().strip()
            mask = df_final_grid['Descrição'].str.lower().str.contains(tb, na=False)
            if 'Cód. ERP' in df_final_grid.columns:
                mask = mask | df_final_grid['Cód. ERP'].astype(str).str.contains(tb, na=False)
            df_filtrado = df_final_grid[mask]
        else:
            df_filtrado = df_final_grid

        col_cfg_l = {
            "codigo": None,
            "Cód. ERP": st.column_config.NumberColumn(disabled=True, format="%d", width=70), 
            "Fornecedor": st.column_config.TextColumn(disabled=True, width=190), 
            "Descrição": st.column_config.TextColumn(disabled=True, width=230),
            "Estoque ERP": st.column_config.NumberColumn("Estoque", disabled=True, format="%d", width=72), 
            "Média (90d)": st.column_config.NumberColumn("Média", disabled=True, format="%.2f", width=72),
            "Qtde Pedida": st.column_config.TextColumn("Qtde", width=85)
        }

        # use_container_width=False: as colunas respeitam a largura definida e não
        # "esticam" para preencher a tela (igual à Visão Fornecedores). Fica bem mais
        # compacto — importante porque as lojas também acessam pelo celular.
        # A chave inclui o filtro: ao trocar de fornecedor o editor reinicia limpo,
        # evitando que um valor digitado "escorregue" para a linha de outro produto.
        grid_editado = st.data_editor(df_filtrado, column_config=col_cfg_l, hide_index=True, use_container_width=False, key=f"grid_loja_{num_loja}_{filtro_forn}")

        # ⚠️ ITEM 3 — alerta de pedido muito acima da média (10x). grid_editado já traz
        # os valores digitados (mesmo antes de salvar), então o alerta atualiza na hora.
        # Só faz sentido onde existe a coluna Média.
        if usa_media:
            df_check = grid_editado.copy()
            df_check["_q"] = df_check["Qtde Pedida"].apply(converter_para_int_seguro)
            df_check["_m"] = pd.to_numeric(df_check["Média (90d)"], errors="coerce").fillna(0.0)
            # só sinaliza quando há média (>0) — produto novo sem histórico não dá pra julgar
            df_outliers = df_check[(df_check["_m"] > 0) & (df_check["_q"] > 10 * df_check["_m"])]
            if not df_outliers.empty:
                st.warning(f"⚠️ **{len(df_outliers)} item(ns) com pedido acima de 10x a média.** Confira se não há erro de digitação:")
                st.dataframe(
                    df_outliers[["Cód. ERP", "Descrição", "Média (90d)", "Qtde Pedida"]].rename(columns={"Qtde Pedida": "Pedido"}),
                    hide_index=True, use_container_width=True,
                    column_config={"Média (90d)": st.column_config.NumberColumn(format="%.2f")},
                )

        # 📝 Campo de Observação Geral da Loja (no final dos pedidos)
        obs_loja = ""
        if usa_obs:
            st.markdown("<div class='no-print'></div>", unsafe_allow_html=True)
            obs_loja = st.text_area(
                "📝 Observação Geral da Loja",
                value=obs_atual,
                key=f"obs_loja_{setor}_{num_loja}",
                placeholder="Ex.: 2 estiletes; 2 grampeadores; 2 sacos de enforca gato...",
                help="Recados gerais desta loja para a separação (itens fora da lista, pedidos especiais etc.)."
            )

        # 🛡️ ITEM 2 — guarda contra perder a digitação (compara original x editado)
        orig_q = df_filtrado["Qtde Pedida"].fillna("").astype(str).str.strip().tolist()
        edit_q = grid_editado["Qtde Pedida"].fillna("").astype(str).str.strip().tolist()
        obs_mudou = usa_obs and (obs_loja or "").strip() != (obs_atual or "").strip()
        guardar_contra_saida((orig_q != edit_q) or obs_mudou)

        df_print_loja = df_filtrado.drop(columns=['codigo'], errors='ignore').fillna('').rename(columns={'Qtde Pedida': 'Pedido'})
        html_table = df_print_loja.to_html(index=False, classes="print-table")
        obs_html = ""
        if usa_obs and (obs_loja or "").strip():
            obs_fmt = (obs_loja or "").strip().replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")
            obs_html = f'<div class="print-obs-loja"><strong>📝 Observação Geral da Loja:</strong><br>{obs_fmt}</div>'
        st.markdown(f'<div class="print-only print-lojas"><h3>🥬 Pedido Oficial — {loja_selecionada}</h3><div class="print-datetime">Emitido em {data_hora_brasilia()}</div>{html_table}{obs_html}</div>', unsafe_allow_html=True)

        # Linha de botões. No Açougue Adriano/Especiais entra o "Sem Pedido Hoje"
        # (menor) entre Salvar e Exportar; nos demais setores a linha fica como antes.
        if usa_sem_pedido:
            c_salvar, c_sem, c_excel, c_print = st.columns([2, 1.5, 2, 1])
        else:
            c_salvar, c_excel, c_print = st.columns([2, 2, 1])
            c_sem = None

        with c_salvar: 
            btn_salvar_loja = st.button("💾 Salvar Pedido", type="primary", use_container_width=True)
        if c_sem is not None:
            with c_sem:
                # Abre a janela de confirmação centralizada (modal). Só no "Sim" é que
                # zera o pedido da loja no dia e dispara o aviso no Telegram.
                if st.button(
                    "🚫 Sem Pedido Hoje",
                    use_container_width=True,
                    help="Informa que esta loja NÃO fará pedido hoje: zera o que estiver lançado e avisa o supervisor no Telegram.",
                ):
                    modal_sem_pedido(setor, num_loja, loja_selecionada)
        with c_excel:
            df_export = grid_editado.drop(columns=['codigo'], errors='ignore')
            st.download_button("📊 Exportar Excel", data=gerar_excel_download(df_export, f"Pedido", obs_rodape=(obs_loja if usa_obs else "")), file_name=f"Pedido_{loja_selecionada}.xlsx", use_container_width=True)
        with c_print: 
            injetar_botao_impressao()

        if btn_salvar_loja:
            with st.spinner("Gravando pedido..."):
                cods_tela = grid_editado["codigo"].tolist()
                if cods_tela:
                    supabase.table("pedidos").delete().eq("setor", setor).eq("loja", num_loja).eq("data_pedido", str(data_brasilia())).in_("codigo_produto", cods_tela).execute()
                    
                    lista_ins = []
                    for _, r in grid_editado.iterrows():
                        q = converter_para_int_seguro(r["Qtde Pedida"])
                        if q > 0: 
                            lista_ins.append({"data_pedido": str(data_brasilia()), "setor": setor, "loja": num_loja, "codigo_produto": int(r["codigo"]), "quantidade": q, "usuario": usuario_atual})
                    
                    if lista_ins: 
                        supabase.table("pedidos").insert(lista_ins).execute()
                        # se a loja tinha declarado "sem pedido" e agora lançou itens, tira a marca
                        if usa_sem_pedido:
                            remover_sem_pedido(setor, num_loja, str(data_brasilia()))

                # 📝 grava a Observação Geral da Loja (embalagem/padaria/confeitaria)
                if usa_obs:
                    salvar_obs_loja(setor, num_loja, str(data_brasilia()), obs_loja, usuario_atual)
            # limpa o estado dos editores (todas as variações de filtro) p/ o guarda desligar
            for _k in [k for k in list(st.session_state.keys()) if str(k).startswith(f"grid_loja_{num_loja}")]:
                st.session_state.pop(_k, None)
            st.session_state.pop(f"obs_loja_{setor}_{num_loja}", None)
            # ✅ sinaliza sucesso → o modal de confirmação aparece após o rerun
            st.session_state[f"pedido_salvo_ok_{setor}_{num_loja}"] = True
            st.rerun()

    # ─────────────────────────────────────────────────────────────────────────
    # ROTA 3 — VISÃO FORNECEDORES (RESUMO EDITÁVEL BANDEIRADO)
    # ─────────────────────────────────────────────────────────────────────────
    elif perfil_navegacao == "Visão Fornecedores (Resumo)":
        st.markdown(f"<div class='no-print'><h2>🚚 Resumo Consolidado por Fornecedor — {setor}</h2></div>", unsafe_allow_html=True)
        st.markdown(f'<div class="print-only"><h3>🚚 Resumo por Fornecedor — {setor}</h3><div class="print-datetime">Emitido em {data_hora_brasilia()}</div></div>', unsafe_allow_html=True)
        
        df_prod = carregar_produtos(setor).copy()
        resp_ped = supabase.table("pedidos").select("codigo_produto, loja, quantidade").eq("setor", setor).eq("data_pedido", str(data_brasilia())).execute()
        
        df_ped = pd.DataFrame(resp_ped.data)

        if df_prod.empty: 
            st.info("Nenhum produto cadastrado.")
            return

        if 'codigo_erp' not in df_prod.columns: 
            df_prod['codigo_erp'] = df_prod['codigo']
            
        df_prod['codigo_erp'] = df_prod['codigo_erp'].fillna(df_prod['codigo']).astype(int)

        codigos_setor = df_prod['codigo'].tolist()
        df_perm_all = buscar_permissoes_setor(supabase, codigos_setor)
        df_prod['descricao'] = df_prod['nome_personalizado'].apply(lambda x: str(x).strip() if pd.notna(x) and str(x).strip() != "" else None).fillna(df_prod['descricao'])

        if not df_ped.empty:
            df_pivot = df_ped.pivot_table(index='codigo_produto', columns='loja', values='quantidade', aggfunc='sum').reset_index()
            for n in range(1, 9):
                if n in df_pivot.columns: 
                    df_pivot = df_pivot.rename(columns={n: f"Loja {n:02d}"})
        else: 
            df_pivot = pd.DataFrame(columns=['codigo_produto'])
            
        df_mestre = pd.merge(df_prod, df_pivot, left_on='codigo', right_on='codigo_produto', how='left')
        
        for l in LOJAS_NOMES:
            if l not in df_mestre.columns: 
                df_mestre[l] = 0.0
            df_mestre[l] = df_mestre[l].fillna(0).astype(int)

        if not df_perm_all.empty:
            df_perm_all['loja_nome'] = df_perm_all['loja'].apply(lambda x: f"Loja {int(x):02d}_perm")
            df_perm_pivot = df_perm_all.pivot_table(index='codigo_produto', columns='loja_nome', values='disponivel', aggfunc='last').reset_index()
        else: 
            df_perm_pivot = pd.DataFrame(columns=['codigo_produto'])

        df_mestre = pd.merge(df_mestre, df_perm_pivot, left_on='codigo', right_on='codigo_produto', how='left')
        
        perm_cols = [f"Loja {n:02d}_perm" for n in range(1, 9)]
        for col in perm_cols:
            if col not in df_mestre.columns: 
                df_mestre[col] = True
                
        df_mestre[perm_cols] = df_mestre[perm_cols].fillna(True)
        mask_active = df_mestre[perm_cols].any(axis=1)
        df_mestre = df_mestre[mask_active]

        df_mestre["TOTAL_NUM"] = df_mestre[LOJAS_NOMES].sum(axis=1)
        df_mestre["TOTAL GERAL"] = df_mestre["TOTAL_NUM"].replace({0: ""})

        for l in LOJAS_NOMES:
            df_mestre[l] = df_mestre[l].replace({0: ""}).astype(str).replace({"0": "", "0.0": "", "nan": ""})
            perm_col = f"{l}_perm"
            if perm_col in df_mestre.columns: 
                df_mestre.loc[df_mestre[perm_col] != True, l] = "-"

        all_edited_frames = []

        usa_iceasa = setor_usa_iceasa(setor)
        usa_erp = setor_usa_erp(setor)
        texto_setor = setor_pedido_texto(setor)   # Matéria Prima/Embalagem/Padaria: sem Total na tela
        eh_mp = setor_eh_materia_prima(setor)
        mapa_obs_mp = {}
        if eh_mp:
            df_extras_mp = carregar_extras(setor, str(data_brasilia()))
            if not df_extras_mp.empty:
                mapa_obs_mp = dict(zip(df_extras_mp["codigo_produto"], df_extras_mp["observacao"]))
        for forn in sorted(df_mestre["fornecedor"].dropna().unique()):
            df_forn_bruto = df_mestre[df_mestre["fornecedor"] == forn]
            lojas_ativas = [l for l in LOJAS_NOMES if not (df_forn_bruto[l] == "-").all()]
            if usa_iceasa:
                cols_cod_f = ["codigo_erp", "codigo_iceasa"]
            elif usa_erp:
                cols_cod_f = ["codigo_erp"]
            else:
                cols_cod_f = []   # Peças Açougue - Manoel: sem código ERP
            if usa_iceasa and "codigo_iceasa" not in df_forn_bruto.columns:
                df_forn_bruto = df_forn_bruto.assign(codigo_iceasa=None)
            cols_total_f = [] if texto_setor else ["TOTAL GERAL"]
            df_forn_view = df_forn_bruto[["codigo"] + cols_cod_f + ["descricao"] + lojas_ativas + cols_total_f].rename(columns={'codigo_erp': 'Cód. ERP', 'codigo_iceasa': 'Cód. Iceasa', 'descricao': 'Produto'}).sort_values(by='Produto')
            if eh_mp:
                df_forn_view["Observação"] = df_forn_view["codigo"].map(mapa_obs_mp).fillna("").astype(str).replace({"None": "", "nan": "", "<NA>": ""})
            
            with st.container(border=True):
                st.markdown(f"<div class='no-print'><h5>Fornecedor: {forn}</h5></div>", unsafe_allow_html=True)
                col_cfg_f = {
                    "codigo": None,
                    "Cód. ERP": st.column_config.NumberColumn(disabled=True, width=80, format="%d"),
                    "Cód. Iceasa": st.column_config.NumberColumn(disabled=True, width=90, format="%d"),
                    "Produto": st.column_config.TextColumn(disabled=True, width=250),
                    "TOTAL GERAL": st.column_config.TextColumn("TOTAL", disabled=True, width=70)
                }
                for l in lojas_ativas: 
                    col_cfg_f[l] = st.column_config.TextColumn(l, width=85, disabled=False)
                if eh_mp:
                    col_cfg_f["Observação"] = st.column_config.TextColumn("Observação:", width=240, help="Observação do item — sai também na exportação Excel.")
                    # limpa "None" preso no estado quando a célula é apagada
                    _estf = st.session_state.get(f"editor_forn_{forn}")
                    if isinstance(_estf, dict) and isinstance(_estf.get("edited_rows"), dict):
                        for _ch in _estf["edited_rows"].values():
                            if "Observação" in _ch and _ch["Observação"] is None:
                                _ch["Observação"] = ""
                    
                edit_df = st.data_editor(df_forn_view, hide_index=True, use_container_width=False, column_config=col_cfg_f, key=f"editor_forn_{forn}")
                
                # Impressão: p/ FLV o código mostrado é o Iceasa (ERP sai), com regra >9000/vazio
                df_forn_print = df_forn_view.drop(columns=['codigo', 'Observação'], errors='ignore').fillna('')
                if usa_iceasa:
                    df_forn_print['Cód. Iceasa'] = df_forn_print['Cód. Iceasa'].apply(iceasa_para_impressao)
                    df_forn_print = df_forn_print.drop(columns=['Cód. ERP'], errors='ignore')
                html_table = df_forn_print.to_html(index=False, classes="print-table")
                st.markdown(f'<div class="print-only print-fornecedores"><h4 class="supplier-header">🚚 Fornecedor: {forn}</h4>{html_table}</div>', unsafe_allow_html=True)
                
                for l in LOJAS_NOMES:
                    if l not in edit_df.columns: 
                        edit_df[l] = "-"
                
                edit_df["fornecedor"] = forn
                all_edited_frames.append(edit_df)

        st.markdown("<div class='no-print'><br></div>", unsafe_allow_html=True)
        
        if all_edited_frames:
            df_forn_editado_full = pd.concat(all_edited_frames, ignore_index=True)
            c_salvar, c_excel, c_print = st.columns([2, 2, 1])
            with c_salvar: 
                btn_salvar_forn = st.button("💾 Salvar Ajustes do Resumo", type="primary", use_container_width=True)
            with c_excel:
                df_export = df_forn_editado_full.drop(columns=['codigo'], errors='ignore')
                st.download_button("📊 Exportar Fornecedores", data=gerar_excel_fornecedores(df_export, f"Fornecedores", sem_total=texto_setor, com_obs=setor_eh_materia_prima(setor)), file_name=f"Resumo_Fornecedores_{setor}.xlsx", use_container_width=True)
            with c_print: 
                injetar_botao_impressao()

            if btn_salvar_forn:
                with st.spinner("Atualizando registros..."):
                    cods = df_forn_editado_full["codigo"].tolist()
                    if cods:
                        for loja_nome in LOJAS_NOMES:
                            n_loja = int(loja_nome.split()[-1])
                            supabase.table("pedidos").delete().eq("setor", setor).eq("loja", n_loja).eq("data_pedido", str(data_brasilia())).in_("codigo_produto", cods).execute()
                            
                            lista_ins = []
                            for _, r in df_forn_editado_full.iterrows():
                                q = converter_para_int_seguro(r[loja_nome])
                                if q > 0: 
                                    lista_ins.append({"data_pedido": str(data_brasilia()), "setor": setor, "loja": n_loja, "codigo_produto": int(r["codigo"]), "quantidade": q, "usuario": usuario_atual})
                            
                            if lista_ins: 
                                supabase.table("pedidos").insert(lista_ins).execute()

                        # 📝 Observação por produto (só Matéria Prima) → separacao_extras
                        if eh_mp:
                            for i in range(0, len(cods), 200):
                                supabase.table("separacao_extras").delete().eq("setor", setor).eq("data_pedido", str(data_brasilia())).in_("codigo_produto", cods[i:i+200]).execute()
                            lista_extras = []
                            for _, r in df_forn_editado_full.iterrows():
                                obs_raw = r.get("Observação")
                                obs_val = str(obs_raw).strip() if pd.notna(obs_raw) and str(obs_raw).strip() != "" else None
                                if obs_val is not None:
                                    lista_extras.append({"codigo_produto": int(r["codigo"]), "data_pedido": str(data_brasilia()), "setor": setor, "preco": None, "observacao": obs_val})
                            for i in range(0, len(lista_extras), 1000):
                                supabase.table("separacao_extras").insert(lista_extras[i:i+1000]).execute()
                st.success("Alterações consolidadas!")
                st.rerun()

    # ─────────────────────────────────────────────────────────────────────────
    # ROTA — PEDIDO POR FORNECEDOR (só FLV Normal/Ofertas) — impressão p/ WhatsApp
    # ─────────────────────────────────────────────────────────────────────────
    elif perfil_navegacao == "Pedido por Fornecedor":
        st.markdown(f"<div class='no-print'><h2>🚚 Pedido por Fornecedor — {setor}</h2></div>", unsafe_allow_html=True)

        if not setor_usa_iceasa(setor):
            st.info("Esta visão é exclusiva do FLV Normal e FLV Ofertas.")
            return

        df_prod = carregar_produtos(setor).copy()
        if df_prod.empty:
            st.warning("Nenhum produto cadastrado neste setor.")
            return
        if 'codigo_iceasa' not in df_prod.columns:
            df_prod['codigo_iceasa'] = None
        df_prod['descricao'] = df_prod['nome_personalizado'].apply(lambda x: str(x).strip() if pd.notna(x) and str(x).strip() != "" else None).fillna(df_prod['descricao'])

        # Pedidos de hoje → quantidade por loja e total por produto
        resp_ped = supabase.table("pedidos").select("codigo_produto, loja, quantidade").eq("setor", setor).eq("data_pedido", str(data_brasilia())).execute()
        df_ped = pd.DataFrame(resp_ped.data)
        qtd_por_loja, total_por_cod = {}, {}
        if not df_ped.empty:
            for cod, grp in df_ped.groupby("codigo_produto"):
                qtd_por_loja[int(cod)] = {int(r["loja"]): int(r["quantidade"]) for _, r in grp.iterrows()}
                total_por_cod[int(cod)] = int(grp["quantidade"].sum())

        # Preços do dia (da Separação)
        df_extras = carregar_extras(setor, str(data_brasilia()))
        preco_por_cod = dict(zip(df_extras["codigo_produto"], df_extras["preco"])) if not df_extras.empty else {}

        desc_por_cod = dict(zip(df_prod["codigo"], df_prod["descricao"]))
        iceasa_por_cod = dict(zip(df_prod["codigo"], df_prod["codigo_iceasa"]))

        # Rótulos únicos p/ o multiselect (desc; acrescenta código se a desc repetir)
        cont_desc = df_prod["descricao"].astype(str).value_counts()
        def _label(cod):
            d = str(desc_por_cod.get(cod, cod))
            if cont_desc.get(d, 0) > 1:
                ice = iceasa_por_cod.get(cod)
                return f"{d} [{int(ice) if pd.notna(ice) else int(cod)}]"
            return d
        label_por_cod = {int(c): _label(int(c)) for c in df_prod["codigo"]}
        cod_por_label = {v: k for k, v in label_por_cod.items()}
        opcoes_labels = sorted(label_por_cod.values(), key=lambda s: s.lower())

        # Config FIXA (fornecedor → produtos), resolvendo Iceasa → produto do setor atual.
        # Não vem do banco: a visão só puxa quantidade/preço já existentes.
        cod_por_iceasa = {}
        for _, r in df_prod.iterrows():
            ice = r["codigo_iceasa"]
            if pd.notna(ice):
                cod_por_iceasa[int(ice)] = int(r["codigo"])
        config = {}
        for forn, iceasas in MAPA_FORNECEDORES_FLV.items():
            cods = [cod_por_iceasa[int(ic)] for ic in iceasas if int(ic) in cod_por_iceasa]
            config[forn] = {"codigos": cods, "linha": forn in FORNECEDORES_LINHA_FLV}

        # Fornecedores extras adicionados nesta sessão (só p/ ajustar a impressão; não persiste)
        extras_key = f"fp_extras_{setor}"
        st.session_state.setdefault(extras_key, [])
        nomes_forn = list(config.keys()) + [n for n in st.session_state[extras_key] if n not in config]

        def _preco(cod):
            v = preco_por_cod.get(cod)
            try:
                return float(v) if v is not None and not (isinstance(v, float) and pd.isna(v)) else 0.0
            except (ValueError, TypeError):
                return 0.0

        # Topo: adicionar fornecedor (sessão) + imprimir. Edições aqui são só p/ a impressão do dia.
        st.markdown("<div class='no-print'>", unsafe_allow_html=True)
        st.caption("As quantidades e preços vêm dos pedidos das lojas e da Separação. Ajustes de nome/produtos aqui valem só para a impressão atual.")
        c_add, c_addbtn = st.columns([3, 1.2])
        with c_add:
            novo_nome = st.text_input("➕ Novo fornecedor:", key=f"fp_novo_{setor}", placeholder="Nome do fornecedor")
        with c_addbtn:
            st.write("")
            if st.button("Adicionar", use_container_width=True, key=f"fp_addbtn_{setor}"):
                nn = str(novo_nome).strip()
                if nn and nn not in nomes_forn:
                    st.session_state[extras_key].append(nn)
                    st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

        # Render dos cards (2 colunas) — edição (só p/ impressão) e montagem do bloco de print
        blocos_print = []
        for i in range(0, len(nomes_forn), 2):
            cols = st.columns(2, gap="small")
            for j, forn in enumerate(nomes_forn[i:i + 2]):
                cfg = config.get(forn, {"codigos": [], "linha": False})
                with cols[j]:
                    with st.container(border=True):
                        st.markdown("<div class='no-print'>", unsafe_allow_html=True)
                        nome_edit = st.text_input("Fornecedor", value=forn, key=f"fp_nome_{setor}_{i+j}", label_visibility="collapsed")
                        default_labels = [label_por_cod[c] for c in cfg["codigos"] if c in label_por_cod]
                        sel_labels = st.multiselect("Produtos", options=opcoes_labels, default=default_labels, key=f"fp_prod_{setor}_{i+j}", label_visibility="collapsed", placeholder="Produtos deste fornecedor")
                        linha_edit = st.checkbox("Layout girado (lojas nas linhas)", value=cfg["linha"], key=f"fp_linha_{setor}_{i+j}")
                        st.markdown("</div>", unsafe_allow_html=True)

                        nome_final = str(nome_edit).strip()
                        cods_sel = [cod_por_label[l] for l in sel_labels if l in cod_por_label]

                        if not cods_sel:
                            st.caption("Sem produtos selecionados.")
                            continue

                        if linha_edit:
                            dados = {"Loja": [f"{n:02d}" for n in range(1, 9)] + ["TOTAL"]}
                            for cod in cods_sel:
                                desc = str(desc_por_cod.get(cod, cod))
                                ice = iceasa_por_cod.get(cod)
                                cab = f"{int(ice) if pd.notna(ice) else cod} - {' '.join(desc.split()[:2])}"
                                dl = qtd_por_loja.get(cod, {})
                                vals = [dl.get(n, 0) for n in range(1, 9)]
                                vals.append(sum(vals))
                                dados[cab] = vals
                            df_card = pd.DataFrame(dados)
                            st.dataframe(df_card, hide_index=True, use_container_width=True)
                            blocos_print.append(f'<h4 class="supplier-header">🛒 {nome_final}</h4>' + df_card.to_html(index=False, classes="print-table"))
                        else:
                            linhas_n, soma_final = [], 0.0
                            for cod in cods_sel:
                                tot = total_por_cod.get(cod, 0)
                                pre = _preco(cod)
                                rtot = tot * pre
                                soma_final += rtot
                                ice = iceasa_por_cod.get(cod)
                                linhas_n.append({
                                    "Cód": iceasa_para_impressao(ice) if pd.notna(ice) else "",
                                    "Produto": str(desc_por_cod.get(cod, cod)),
                                    "Total": tot if tot > 0 else "",
                                    "R$ Preço": (f"R$ {pre:.2f}".replace(".", ",")) if pre > 0 else "",
                                    "R$ Total": (f"R$ {rtot:.2f}".replace(".", ",")) if rtot > 0 else "",
                                })
                            df_card = pd.DataFrame(linhas_n)
                            st.dataframe(df_card, hide_index=True, use_container_width=True)
                            tf_str = f"R$ {soma_final:.2f}".replace(".", ",")
                            st.markdown(f"<div class='no-print' style='text-align:right;font-weight:700;color:#1f7a3d;'>Total Final: {tf_str}</div>", unsafe_allow_html=True)
                            blocos_print.append(
                                f'<h4 class="supplier-header">🛒 {nome_final}</h4>'
                                + df_card.to_html(index=False, classes="print-table")
                                + f'<div style="text-align:right;font-weight:bold;margin:2px 0 10px 0;">Total Final: {tf_str}</div>'
                            )

        # Bloco de impressão consolidado (fora das colunas, p/ não ser escondido na impressão)
        if blocos_print:
            corpo = "".join(f'<div style="page-break-inside:avoid;">{b}</div>' for b in blocos_print)
            st.markdown(
                f'<div class="print-only"><h3>🚚 Pedido por Fornecedor — {setor}</h3>'
                f'<div class="print-datetime">Emitido em {data_hora_brasilia()}</div>{corpo}</div>',
                unsafe_allow_html=True,
            )

        st.markdown("<div class='no-print'><br></div>", unsafe_allow_html=True)
        c_print, _ = st.columns([1, 4])
        with c_print:
            injetar_botao_impressao()

    # ─────────────────────────────────────────────────────────────────────────
    # ROTA 4 — CATÁLOGO DE PRODUTOS
    # ─────────────────────────────────────────────────────────────────────────
    elif perfil_navegacao == "Catálogo de Produtos":
        st.markdown(f"<div class='no-print'><h2>🗂️ Gestão de Catálogo e Permissões por Loja — {setor}</h2></div>", unsafe_allow_html=True)

        # tela de edição → sempre dados frescos (evita editar em cima de permissão velha)
        buscar_permissoes_setor.clear()

        resp_prod = supabase.table("produtos").select("*").eq("setor", setor).execute()
        df_prod = pd.DataFrame(resp_prod.data)

        if df_prod.empty: 
            df_prod = pd.DataFrame(columns=['codigo', 'codigo_erp', 'descricao', 'fornecedor', 'nome_personalizado', 'setor', 'ativo'])
        else:
            if 'codigo_erp' not in df_prod.columns: 
                df_prod['codigo_erp'] = df_prod['codigo']
            df_prod['codigo_erp'] = df_prod['codigo_erp'].fillna(df_prod['codigo']).astype(int)

        codigos_setor = df_prod['codigo'].tolist() if not df_prod.empty else []
        df_perm = buscar_permissoes_setor(supabase, codigos_setor)

        if not df_perm.empty:
            df_perm['loja_nome'] = df_perm['loja'].apply(lambda x: f"Loja {int(x):02d}")
            df_perm_pivot = df_perm.pivot_table(index='codigo_produto', columns='loja_nome', values='disponivel', aggfunc='last').reset_index()
        else: 
            df_perm_pivot = pd.DataFrame(columns=['codigo_produto'] + LOJAS_NOMES)

        if not df_prod.empty: 
            df_cat_completo = pd.merge(df_prod, df_perm_pivot, left_on='codigo', right_on='codigo_produto', how='left').drop(columns=['codigo_produto'], errors='ignore')
        else: 
            df_cat_completo = pd.DataFrame(columns=['codigo', 'codigo_erp', 'descricao', 'fornecedor', 'nome_personalizado'] + LOJAS_NOMES)
        
        for l in LOJAS_NOMES: 
            if l not in df_cat_completo.columns: 
                df_cat_completo[l] = True
            else: 
                df_cat_completo[l] = df_cat_completo[l].fillna(True).astype(bool)
        
        if 'nome_personalizado' not in df_cat_completo.columns: 
            df_cat_completo['nome_personalizado'] = ""
        else: 
            df_cat_completo['nome_personalizado'] = df_cat_completo['nome_personalizado'].fillna("")

        usa_iceasa = setor_usa_iceasa(setor)
        usa_erp = setor_usa_erp(setor)
        if usa_iceasa and 'codigo_iceasa' not in df_cat_completo.columns:
            df_cat_completo['codigo_iceasa'] = None

        if not df_cat_completo.empty: 
            df_cat_completo = df_cat_completo.sort_values(by=['fornecedor', 'descricao']).reset_index(drop=True)

        col_cfg_c = {
            "codigo": None,
            "descricao": st.column_config.TextColumn("Nome Prime", width=180), 
            "nome_personalizado": st.column_config.TextColumn("Nome Manual", width=160),
            "fornecedor": st.column_config.TextColumn("Fornecedor", width=130)
        }
        if usa_erp:
            col_cfg_c["codigo_erp"] = st.column_config.NumberColumn("Cód. ERP", format="%d", width=80)
        if usa_iceasa:
            col_cfg_c["codigo_iceasa"] = st.column_config.NumberColumn("Cód. Iceasa", format="%d", width=90)
        for l in LOJAS_NOMES: 
            col_cfg_c[l] = st.column_config.CheckboxColumn(l)

        cols_base = ["fornecedor", "codigo"]
        if usa_erp:
            cols_base.append("codigo_erp")
        if usa_iceasa:
            cols_base.append("codigo_iceasa")
        cols_exibicao = cols_base + ["descricao", "nome_personalizado"] + LOJAS_NOMES
        edited_cat = st.data_editor(df_cat_completo[cols_exibicao], use_container_width=True, hide_index=True, column_config=col_cfg_c, num_rows="dynamic", key="catalogo_editor")

        html_table = df_cat_completo[cols_exibicao].drop(columns=['codigo'], errors='ignore').fillna('').to_html(index=False, classes="print-table")
        st.markdown(f'<div class="print-only"><h3>🗂️ Catálogo Geral — {setor}</h3><div class="print-datetime">Emitido em {data_hora_brasilia()}</div>{html_table}</div>', unsafe_allow_html=True)

        st.markdown("<div class='no-print'><br></div>", unsafe_allow_html=True)
        col_btn_salvar, col_btn_erp = st.columns(2)
        
        with col_btn_salvar: 
            btn_salvar = st.button("💾 Salvar Matriz do Catálogo", type="primary", use_container_width=True)
        with col_btn_erp: 
            btn_puxar_erp = st.button("📥 Puxar Nomes do ERP", use_container_width=True)

        if btn_salvar:
            state = st.session_state.get("catalogo_editor")
            with st.spinner("Automação Duplo-Código processando..."):
                try:
                    resp_all = supabase.table("produtos").select("codigo").execute()
                    codigos_globais = [p["codigo"] for p in resp_all.data] if resp_all.data else []
                    codigos_conhecidos = set(df_cat_completo['codigo'].dropna().astype(int).tolist()) if not df_cat_completo.empty else set()
                    
                    mapa_novos_idx = {}  # índice da linha nova → código (PK) gerado

                    if state and state.get("deleted_rows"):
                        for idx in state["deleted_rows"]:
                            cod_p = int(df_cat_completo.iloc[idx]["codigo"])
                            supabase.table("produtos_lojas").delete().eq("codigo_produto", cod_p).execute()
                            supabase.table("pedidos").delete().eq("codigo_produto", cod_p).execute()
                            supabase.table("medias_90d").delete().eq("codigo_produto", cod_p).execute()
                            supabase.table("produtos").delete().eq("codigo", cod_p).execute()
                            if cod_p in codigos_globais: codigos_globais.remove(cod_p)
                            if cod_p in codigos_conhecidos: codigos_conhecidos.remove(cod_p)

                    if state and state.get("edited_rows"):
                        for idx_str, changes in state["edited_rows"].items():
                            idx = int(idx_str)
                            cod_p_original = int(df_cat_completo.iloc[idx]["codigo"])
                            prod_changes = {}
                            if "descricao" in changes: 
                                prod_changes["descricao"] = str(changes["descricao"])
                            if "fornecedor" in changes: 
                                prod_changes["fornecedor"] = str(changes["fornecedor"])
                            if "nome_personalizado" in changes: 
                                val_np = changes["nome_personalizado"]
                                prod_changes["nome_personalizado"] = str(val_np).strip() if pd.notna(val_np) and str(val_np).strip() != "" else None
                            if "codigo_erp" in changes:
                                try: prod_changes["codigo_erp"] = int(changes["codigo_erp"])
                                except: pass
                            if "codigo_iceasa" in changes:
                                val_ice = changes["codigo_iceasa"]
                                try:
                                    prod_changes["codigo_iceasa"] = int(val_ice) if pd.notna(val_ice) and str(val_ice).strip() != "" else None
                                except (ValueError, TypeError):
                                    prod_changes["codigo_iceasa"] = None
                            if prod_changes: 
                                supabase.table("produtos").update(prod_changes).eq("codigo", cod_p_original).execute()

                    for idx, row in edited_cat.iterrows():
                        c_pk = row.get("codigo")
                        # Linha já existente (tem PK) não é inserida aqui.
                        if pd.notna(c_pk) and str(c_pk).strip() != "":
                            continue

                        c_erp = row.get("codigo_erp")
                        tem_erp = usa_erp and pd.notna(c_erp) and str(c_erp).strip() != ""
                        forn_add = str(row.get("fornecedor", "Box")).strip()
                        desc_add = str(row.get("descricao", "")).strip()

                        # Sem ERP e sem descrição → linha vazia, ignora.
                        if not tem_erp and (desc_add == "" or desc_add.lower() == "nan"):
                            continue
                        if desc_add == "" or desc_add.lower() == "nan":
                            desc_add = "Novo Produto"

                        np_add = str(row.get("nome_personalizado", "")).strip() if pd.notna(row.get("nome_personalizado")) and str(row.get("nome_personalizado")).strip() != "" else None
                        ice_add = None
                        if usa_iceasa:
                            v_ice = row.get("codigo_iceasa")
                            try:
                                ice_add = int(v_ice) if pd.notna(v_ice) and str(v_ice).strip() != "" else None
                            except (ValueError, TypeError):
                                ice_add = None

                        if tem_erp:
                            cod_erp_digitado = int(float(c_erp))
                            cod_final = cod_erp_digitado
                            if cod_final in codigos_globais:
                                base_str = str(cod_final)
                                for i in range(1, 100):
                                    tent = int(f"{base_str}{i:02d}")
                                    if tent not in codigos_globais:
                                        cod_final = tent
                                        break
                                st.toast(f"🤖 Gerado código interno invisível {cod_final} para o item {cod_erp_digitado} do ERP.", icon="✨")
                        else:
                            # Setor sem ERP (ex.: Peças Açougue - Manoel): gera um PK novo sequencial.
                            cod_erp_digitado = None
                            cod_final = (max(codigos_globais) + 1) if codigos_globais else 1

                        mapa_novos_idx[idx] = cod_final
                        supabase.table("produtos").insert({
                            "codigo": cod_final, "codigo_erp": cod_erp_digitado,
                            "codigo_iceasa": ice_add,
                            "descricao": desc_add, "fornecedor": forn_add, 
                            "nome_personalizado": np_add, "setor": setor, "ativo": True
                        }).execute()
                        codigos_globais.append(cod_final); codigos_conhecidos.add(cod_final)

                    lista_perms_geral = []
                    codigos_processados_perms = set()
                    
                    for idx, row in edited_cat.iterrows():
                        c_pk = row.get("codigo")
                        if pd.notna(c_pk) and str(c_pk).strip() != "":
                            c_final = int(c_pk)
                        else:
                            c_final = mapa_novos_idx.get(idx, None)
                        if not c_final: continue
                        
                        codigos_processados_perms.add(c_final)
                        for num_loja in range(1, 9):
                            val_loja = bool(row.get(f"Loja {num_loja:02d}", True))
                            lista_perms_geral.append({"codigo_produto": c_final, "loja": num_loja, "disponivel": val_loja})

                    codigos_lista = list(codigos_processados_perms)
                    if codigos_lista:
                        for i in range(0, len(codigos_lista), 200): supabase.table("produtos_lojas").delete().in_("codigo_produto", codigos_lista[i:i+200]).execute()
                        for i in range(0, len(lista_perms_geral), 1000): supabase.table("produtos_lojas").insert(lista_perms_geral[i:i+1000]).execute()

                    st.success("✅ Automação concluída!"); st.cache_data.clear(); time.sleep(1.5); st.rerun()
                except Exception as e: st.error(f"⚠️ Erro processando: {e}")

        if btn_puxar_erp:
            with st.spinner("Buscando nomes oficias usando Cód. ERP..."):
                try:
                    cods_erp = [int(c) for c in edited_cat["codigo_erp"].tolist() if pd.notna(c) and str(c).strip() != ""]
                    if not cods_erp: st.warning("Nenhum código encontrado.")
                    else:
                        cods_str = ", ".join(map(str, set(cods_erp)))
                        query_nomes = f"SELECT cod, descricao FROM python_ajuste_cadastro WHERE cod IN ({cods_str})"
                        df_nomes = conn_pg.query(query_nomes, ttl=0)

                        if not df_nomes.empty:
                            for _, row in df_nomes.iterrows():
                                cod_oficial = int(row["cod"])
                                desc_erp = str(row["descricao"])
                                supabase.table("produtos").update({"descricao": desc_erp}).eq("codigo_erp", cod_oficial).execute()
                            st.success("✅ Nomes atualizados em todos os fornecedores!"); st.cache_data.clear(); time.sleep(1); st.rerun()
                        else: st.info("Nenhum nome encontrado.")
                except Exception as e:
                    if "No database configured" in str(e) or "missing" in str(e).lower(): st.error("⚠️ Aviso: Credenciais do PostgreSQL não configuradas ou inacessíveis.")
                    else: st.error(f"⚠️ Erro ao buscar nomes no banco ERP: {e}")
